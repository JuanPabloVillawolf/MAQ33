# -*- coding: utf-8 -*-
"""
FLEXCON_Extractor_Universal_v3.py
==================================
App Streamlit — Extractor de documentos FLEXCON PDF → Excel

Uso:  streamlit run flexcon_extractor_universal_v3.py
"""

import re, io, os, warnings
import pandas as pd
import pdfplumber
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ══════════════════════════════════════════════════════════════════════════════
#  ZONAS DE COLUMNA  (posición X en puntos PDF)
# ══════════════════════════════════════════════════════════════════════════════

INV_COLS = {
    "LINE":        (0,    44),
    "B_P":         (44,   73),
    "ITEM":        (73,  145),
    "DESCRIPTION": (145, 369),
    "ORIGIN":      (369, 427),
    "QUANTITY":    (427, 492),
    "UOM":         (492, 534),
    "QTY_M":       (534, 585),
    # BOXES  585-635 → manuscrito, se ignora
    # WEIGHT 635-712 → manuscrito, se ignora
    "VALUE":       (712, 9999),
}

NI_COLS = {
    "LINE":        (0,   100),   # 100 para capturar "THINNER" a x0=104 aprox
    "ITEM_NUMBER": (100, 230),
    "DESCRIPTION": (230, 483),
    "UOM":         (483, 515),
    "ORIGIN":      (515, 545),
    "QTY":         (545, 605),
    "LOT":         (605, 655),
    "VALUE":       (655, 710),
    # BOXES/WEIGHT → manuscrito
}

ORIGINS  = {"USA","JPN","NLD","MEX","CHN","KOR","TWN","DEU"}
LOT_RE   = re.compile(r'^[A-Za-z][Oo0-9]{7,}$')
VALUE_RE = re.compile(r'^\d[\d,]*\.\d{2}$')
INT_RE   = re.compile(r'^\d[\d,]*$')
# Palabras de formulario a ignorar en NI
FORM_WORDS = {"shipped","received","shipping","notes","date","total","value","boxes","by"}


# ══════════════════════════════════════════════════════════════════════════════
#  UTILIDADES
# ══════════════════════════════════════════════════════════════════════════════

def col_of(x: float, zones: dict):
    for col, (lo, hi) in zones.items():
        if lo <= x < hi:
            return col
    return None

def is_garbage(text: str) -> bool:
    """Descarta artefactos de texto manuscrito (^, _, /, *, ¿, etc.)."""
    t = text.strip()
    if not t:
        return True
    alphanum = sum(1 for c in t if c.isalnum() or c in '$.,')
    return alphanum / max(len(t), 1) < 0.45

def clean_num(s: str) -> str:
    return s.replace(',', '').replace('$', '').strip()

def fix_decimal(s: str) -> str:
    """Convierte '4,754,90' (OCR) → '4754.90'. Reemplaza la última coma por punto."""
    s = s.strip().replace('$', '')
    # Si ya tiene punto decimal, limpiar comas y devolver
    if '.' in s:
        return s.replace(',', '')
    # Sin punto: buscar si parece un decimal con coma al final (e.g. 4,754,90)
    parts = s.split(',')
    if len(parts) >= 2 and len(parts[-1]) == 2:
        return parts[-1].join([''.join(parts[:-1]), '']).replace(',', '') \
               if False else (''.join(parts[:-1]).replace(',','') + '.' + parts[-1])
    return s.replace(',', '')


def parse_quantity(s: str) -> str:
    """Interpreta cantidades con OCR defectuoso.
    '2.720' -> '2720', '300.000' -> '300000', '4,000' -> '4000'
    """
    s = s.strip().replace(',', '')
    # Si tiene punto seguido de MAS de 2 dígitos -> el punto es separador de miles
    if re.match(r'^\d+\.\d{3,}$', s):
        return s.replace('.', '')
    # Entero normal
    if re.match(r'^\d+$', s):
        return s
    return ''

def safe_float(val):
    try:
        if val in (None, '', 'N/A', 'N/D') or (isinstance(val, float) and pd.isna(val)):
            return None
        s = fix_decimal(str(val))
        return float(s)
    except:
        return None

def safe_int(val):
    try:
        if val in (None, '', 'N/A', 'N/D') or (isinstance(val, float) and pd.isna(val)):
            return None
        return int(clean_num(str(val)))
    except:
        return None

def page_text(page) -> str:
    try:
        return page.extract_text(x_tolerance=3, y_tolerance=3) or ""
    except:
        return page.extract_text() or ""

def is_non_inv_page(page) -> bool:
    t = page_text(page).lower()
    return any(m in t for m in
               ["non-inventory transfer","non-lnventory","requestor name",
                "ship to address","m27009","seip - flexcon"])

def get_words(page) -> list:
    try:
        return page.extract_words(x_tolerance=3, y_tolerance=3)
    except:
        return page.extract_words()


# ══════════════════════════════════════════════════════════════════════════════
#  EXTRACCIÓN — INV TRANSFER
# ══════════════════════════════════════════════════════════════════════════════

def extract_inventory_data(pdf_bytes: bytes) -> tuple[list, dict]:
    records  = []
    doc_meta = {}

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            if is_non_inv_page(page):
                continue

            # Metadata del documento (se actualiza en cada página para capturar Total Value)
            txt = page_text(page)
            if True:
                m = re.search(r'\b(P\d{7,})\b', txt)
                if m: doc_meta['doc_number'] = m.group(1)
                m = re.search(r'Transfer Date[:\s]+(\d{1,2}/\d{1,2}/\d{2,4})', txt, re.I)
                if m: doc_meta['transfer_date'] = m.group(1)
                else:
                    m = re.search(r'\b(\d{1,2}/\d{1,2}/\d{2,4})\b', txt)
                    if m: doc_meta['transfer_date'] = m.group(1)
                m = re.search(r'Total\s+Value[:\s]+\$?([\d,]+\.\d{2})', txt, re.I)
                if m: doc_meta['total_value'] = float(clean_num(m.group(1)))

            words = get_words(page)

            # ── Paso 1: Encontrar todos los words con sus coords reales ──────
            # Guardamos (x0, top, text) por palabra no basura
            clean_words = []
            for w in words:
                t = w['text'].strip()
                if t and not is_garbage(t):
                    clean_words.append({'x0': w['x0'], 'top': w['top'], 'text': t})

            # ── Paso 2: Localizar filas ancla (palabras en zona LINE, número 1-99) ──
            anchor_list = []  # [(actual_top, line_num)]
            for cw in clean_words:
                if col_of(cw['x0'], INV_COLS) == 'LINE':
                    if re.match(r'^\d{1,3}$', cw['text']) and 1 <= int(cw['text']) <= 99:
                        anchor_list.append((cw['top'], int(cw['text'])))

            # ── Paso 3: Para cada ancla, recopilar campos por ventana Y ──────
            for actual_top, line_num in anchor_list:

                def collect(zones_keys, dy_max):
                    """Recoge textos de ciertas zonas dentro de ±dy_max del ancla."""
                    result = {k: [] for k in zones_keys}
                    for cw in clean_words:
                        dy = abs(cw['top'] - actual_top)
                        if dy > dy_max:
                            continue
                        col = col_of(cw['x0'], INV_COLS)
                        if col in zones_keys:
                            result[col].append((cw['top'], cw['text']))
                    return result

                def collect_below(zones_keys, dy_min, dy_max):
                    """Recoge textos debajo del ancla (para LOT)."""
                    result = {k: [] for k in zones_keys}
                    for cw in clean_words:
                        dy = cw['top'] - actual_top
                        if dy_min <= dy <= dy_max:
                            col = col_of(cw['x0'], INV_COLS)
                            if col in zones_keys:
                                result[col].append(cw['text'])
                    return result

                # Campos de la misma fila impresa: ±6px
                same_row = collect(['B_P','ITEM','DESCRIPTION','ORIGIN','QUANTITY'], 6)
                # Campos flotantes (UOM, QTY_M, VALUE): ±18px
                float_row = collect(['UOM','QTY_M','VALUE'], 18)
                # LOT: debajo entre +6 y +20px, en zona DESCRIPTION
                lot_data  = collect_below(['DESCRIPTION'], 6, 20)

                # ── Construir campos ─────────────────────────────────────────
                bp   = ' '.join(t for _, t in same_row['B_P'])
                item = ' '.join(t for _, t in same_row['ITEM'])
                desc = ' '.join(t for _, t in sorted(same_row['DESCRIPTION']))

                # ORIGIN: primer token válido en ±6px
                origin = next((t for _, t in same_row['ORIGIN']
                                if t.upper() in ORIGINS), '')

                # QUANTITY: primer número entero en ±6px (maneja OCR '2.720' -> 2720)
                qty_tokens = [parse_quantity(t) for _, t in same_row['QUANTITY']
                              if parse_quantity(t)]
                quantity   = qty_tokens[0] if qty_tokens else ''

                # UOM: primera palabra en ±18px que sea unidad
                uom_tokens = [t.upper() for _, t in float_row['UOM']
                              if not is_garbage(t) and len(t) <= 5]
                uom        = uom_tokens[0] if uom_tokens else ''

                # QTY_M: primer entero en ±18px
                qtym_tokens = [clean_num(t) for _, t in float_row['QTY_M']
                               if INT_RE.match(clean_num(t))]
                qty_m       = qtym_tokens[0] if qtym_tokens else ''

                # VALUE: el ÚLTIMO decimal válido en ±18px
                # (el último suele ser el más a la derecha = VALUE real)
                val_candidates = []
                for _, t in float_row['VALUE']:
                    cv = fix_decimal(t)
                    if re.match(r'^\d+\.\d{2}$', cv):
                        val_candidates.append(cv)
                value = val_candidates[-1] if val_candidates else ''

                # LOT: primer token que coincida con el patrón
                lot_candidates = [t for t in lot_data['DESCRIPTION']
                                  if LOT_RE.match(t.replace('O','0').replace('o','0'))]
                if lot_candidates:
                    desc = (desc + ' ' + lot_candidates[0]).strip()

                records.append({
                    'LINE':            line_num,
                    'B/P':             bp,
                    'ITEM':            item,
                    'DESCRIPTION/LOT': desc,
                    'ORIGIN':          origin.upper() if origin else '',
                    'QUANTITY':        quantity,
                    'UOM':             uom,
                    'QTY(M)':          qty_m,
                    'BOXES':           'N/D',
                    'WEIGHT(KG)':      'N/D',
                    'VALUE':           value,
                })

    return records, doc_meta


# ══════════════════════════════════════════════════════════════════════════════
#  EXTRACCIÓN — NON-INVENTORY TRANSFER
# ══════════════════════════════════════════════════════════════════════════════

def extract_non_inventory_data(pdf_bytes: bytes) -> tuple[list, dict]:
    records  = []
    doc_meta = {}

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            if not is_non_inv_page(page):
                continue

            # Metadata
            if not doc_meta:
                txt = page_text(page)
                for pat, key in [
                    (r'Requestor Name[;:\s]+(.+?)(?:\s{2,}|Shipment)', 'requestor'),
                    (r'Shipment Date[;:\s]+(\d{1,2}/\d{1,2}/\d{2,4})',  'shipment_date'),
                    (r'Ship To Address[;:\s]+(.+)',                       'ship_to'),
                    (r'From Location[;:\s]+(.+?)(?:\s{2,}|Transfer)',    'from_location'),
                ]:
                    m = re.search(pat, txt, re.I)
                    if m: doc_meta[key] = m.group(1).strip()
                # TOTAL VALUE puede estar en línea separada, buscar patrón flexible
                m = re.search(r'TOTAL\s+VALUE[^\d$]*\$?\s*([\d,]+\.\d{2})', txt, re.I | re.DOTALL)
                if not m:
                    # Fallback: buscar cualquier $X,XXX.XX al final del texto
                    all_vals = re.findall(r'\$([\d,]+\.\d{2})', txt)
                    if all_vals:
                        doc_meta['total_value'] = float(clean_num(all_vals[-1]))
                else:
                    doc_meta['total_value'] = float(clean_num(m.group(1)))

            words = get_words(page)
            clean_words = [
                {'x0': w['x0'], 'top': w['top'], 'text': w['text'].strip()}
                for w in words
                if w['text'].strip() and not is_garbage(w['text'])
            ]

            # ── Localizar filas ancla (LINE number en zona LINE, 1-20) ────────
            anchor_list = []
            for cw in clean_words:
                if col_of(cw['x0'], NI_COLS) == 'LINE':
                    if re.match(r'^\d{1,2}$', cw['text']) and 1 <= int(cw['text']) <= 20:
                        anchor_list.append((cw['top'], int(cw['text'])))

            if not anchor_list:
                continue

            # ── Ordenar anclas y calcular límite inferior de cada línea ───────
            anchor_list.sort()
            # El límite superior de la línea N es el top de la línea N+1 menos un margen
            def y_limit(idx):
                if idx + 1 < len(anchor_list):
                    return anchor_list[idx + 1][0] - 2
                return anchor_list[idx][0] + 22  # última línea

            # ── Para cada ancla, recolectar datos dentro de [anchor_top, limit] ─
            for i, (actual_top, line_num) in enumerate(anchor_list):
                lo = actual_top - 3     # pequeño margen arriba (algunos datos vienen 1-2px antes)
                hi = y_limit(i)

                combined = {}
                for cw in clean_words:
                    if lo <= cw['top'] <= hi:
                        col = col_of(cw['x0'], NI_COLS)
                        if col and col != 'LINE':
                            combined.setdefault(col, []).append(cw['text'])

                item_num = ' '.join(combined.get('ITEM_NUMBER', []))
                desc     = ' '.join(combined.get('DESCRIPTION', []))
                uom      = ' '.join(combined.get('UOM', [])).upper()
                origin   = next((t for t in combined.get('ORIGIN', [])
                                  if t.upper() in ORIGINS), '')

                # QTY: primer entero válido (maneja OCR con punto como separador de miles)
                qty_raw = [parse_quantity(t) for t in combined.get('QTY', [])
                           if parse_quantity(t)]
                quantity = qty_raw[0] if qty_raw else ''

                lot = next((t for t in combined.get('LOT', [])
                            if t.upper() in ('N/A', 'NA')), 'N/A')

                # VALUE: primer decimal válido en zona VALUE
                val_candidates = []
                for t in combined.get('VALUE', []):
                    cv = fix_decimal(t.replace('$','').replace(' ',''))
                    if re.match(r'^\d+\.\d{2}$', cv):
                        val_candidates.append(cv)
                value = val_candidates[0] if val_candidates else ''

                # Saltar filas vacías o de palabras del formulario
                item_lower = item_num.lower()
                if not item_num.strip() and not desc.strip():
                    continue
                if any(fw in item_lower for fw in FORM_WORDS):
                    continue

                records.append({
                    'LINE':          line_num,
                    'ITEM NUMBER':   item_num,
                    'DESCRIPTION':   desc,
                    'U/M':           uom,
                    'ORIGIN':        origin.upper() if origin else '',
                    'QTY REQUESTED': quantity,
                    'LOT NUMBER':    lot,
                    'VALUE':         value,
                    'BOXES':         'N/D',
                    'WEIGHT(KG)':    'N/D',
                })

    return records, doc_meta


# ══════════════════════════════════════════════════════════════════════════════
#  CONSTRUCCIÓN DEL EXCEL  (en memoria → bytes)
# ══════════════════════════════════════════════════════════════════════════════

def _thick(): return Side(style='medium', color='404040')
def _thin():  return Side(style='thin',   color='C0C0C0')

def _header_style(ws, row_num, hex_color):
    t = _thick()
    f = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
    for cell in ws[row_num]:
        cell.fill      = f
        cell.font      = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        cell.border    = Border(left=t, right=t, top=t, bottom=t)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[row_num].height = 30

def _data_style(ws, r_start, r_end, alt_hex):
    tb = Border(left=_thin(), right=_thin(), top=_thin(), bottom=_thin())
    af = PatternFill(start_color=alt_hex, end_color=alt_hex, fill_type='solid')
    for ri in range(r_start, r_end + 1):
        for cell in ws[ri]:
            cell.border    = tb
            cell.font      = Font(name='Arial', size=9)
            cell.alignment = Alignment(vertical='center')
            if ri % 2 == 0:
                cell.fill = af

def _total_row(ws, r_start, r_end, headers, money_col, sum_cols, total_hex):
    t  = _thick()
    tf = PatternFill(start_color=total_hex, end_color=total_hex, fill_type='solid')
    tr = r_end + 1
    ws.append(['TOTAL'] + [''] * (len(headers) - 1))
    for col in sum_cols:
        if col in headers:
            cl = headers[col]
            ws[f'{cl}{tr}'] = f'=IFERROR(SUM({cl}{r_start}:{cl}{r_end}),"")'
            ws[f'{cl}{tr}'].number_format = '#,##0'
    if money_col in headers:
        cl = headers[money_col]
        ws[f'{cl}{tr}'] = f'=IFERROR(SUM({cl}{r_start}:{cl}{r_end}),"")'
        ws[f'{cl}{tr}'].number_format = '$#,##0.00'
    for cell in ws[tr]:
        cell.fill   = tf
        cell.font   = Font(bold=True, name='Arial', size=10)
        cell.border = Border(left=t, right=t, top=t, bottom=t)


def build_excel(inv_records, inv_meta, noninv_records, noninv_meta) -> bytes:
    wb = Workbook()

    # ── Hoja 1: INV Transfer ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'INV Transfer'

    if inv_records:
        df1 = pd.DataFrame(inv_records)
        df1['LINE']     = df1['LINE'].apply(safe_int)
        df1['QUANTITY'] = df1['QUANTITY'].apply(safe_int)
        df1['QTY(M)']   = df1['QTY(M)'].apply(safe_int)
        df1['VALUE']    = df1['VALUE'].apply(safe_float)
        df1.sort_values('LINE', inplace=True, ignore_index=True)

        meta = (f"Doc: {inv_meta.get('doc_number','')}   "
                f"Fecha: {inv_meta.get('transfer_date','')}   "
                f"Total declarado: ${inv_meta.get('total_value',0) or 0:,.2f}")
        ws1.append([meta])
        nc = len(df1.columns)
        ws1.merge_cells(f"A1:{get_column_letter(nc)}1")
        ws1['A1'].font      = Font(bold=True, italic=True, name='Arial', size=9, color='404040')
        ws1['A1'].alignment = Alignment(horizontal='left')

        ws1.append(list(df1.columns))
        _header_style(ws1, 2, '1F4E78')

        for _, row in df1.iterrows():
            ws1.append(list(row))

        ds, de = 3, 2 + len(df1)
        _data_style(ws1, ds, de, 'EEF3F8')
        hdr = {c.value: c.column_letter for c in ws1[2]}

        for ri in range(ds, de + 1):
            for col in ['QUANTITY', 'QTY(M)']:
                if col in hdr:
                    ws1[f"{hdr[col]}{ri}"].number_format = '#,##0'
            if 'VALUE' in hdr:
                ws1[f"{hdr['VALUE']}{ri}"].number_format = '$#,##0.00'

        _total_row(ws1, ds, de, hdr, 'VALUE', ['QUANTITY', 'QTY(M)'], 'C9D8EC')

        widths = {'A':6,'B':7,'C':16,'D':52,'E':9,'F':12,'G':7,'H':10,'I':7,'J':10,'K':13}
        for cl, w in widths.items():
            ws1.column_dimensions[cl].width = w
        ws1.freeze_panes = 'A3'
    else:
        ws1['A1'] = '⚠️ No se encontraron registros INV TRANSFER'

    # ── Hoja 2: Non-Inv Transfer ─────────────────────────────────────────────
    ws2 = wb.create_sheet('Non-Inv Transfer')

    if noninv_records:
        df2 = pd.DataFrame(noninv_records)
        df2['LINE']          = df2['LINE'].apply(safe_int)
        df2['QTY REQUESTED'] = df2['QTY REQUESTED'].apply(safe_int)
        df2['VALUE']         = df2['VALUE'].apply(safe_float)
        df2.sort_values('LINE', inplace=True, ignore_index=True)

        meta2 = (f"Requestor: {noninv_meta.get('requestor','')}   "
                 f"From: {noninv_meta.get('from_location','')}   "
                 f"Ship To: {noninv_meta.get('ship_to','')}   "
                 f"Fecha: {noninv_meta.get('shipment_date','')}   "
                 f"Total declarado: ${noninv_meta.get('total_value',0) or 0:,.2f}")
        ws2.append([meta2])
        nc2 = len(df2.columns)
        ws2.merge_cells(f"A1:{get_column_letter(nc2)}1")
        ws2['A1'].font      = Font(bold=True, italic=True, name='Arial', size=9, color='404040')
        ws2['A1'].alignment = Alignment(horizontal='left')

        ws2.append(list(df2.columns))
        _header_style(ws2, 2, '1D6A3A')

        for _, row in df2.iterrows():
            ws2.append(list(row))

        ds2, de2 = 3, 2 + len(df2)
        _data_style(ws2, ds2, de2, 'EDF7EF')
        hdr2 = {c.value: c.column_letter for c in ws2[2]}

        for ri in range(ds2, de2 + 1):
            if 'QTY REQUESTED' in hdr2:
                ws2[f"{hdr2['QTY REQUESTED']}{ri}"].number_format = '#,##0'
            if 'VALUE' in hdr2:
                ws2[f"{hdr2['VALUE']}{ri}"].number_format = '$#,##0.00'

        _total_row(ws2, ds2, de2, hdr2, 'VALUE', ['QTY REQUESTED'], 'C6E8CC')

        widths2 = {'A':6,'B':16,'C':48,'D':7,'E':9,'F':14,'G':10,'H':12,'I':7,'J':10}
        for cl, w in widths2.items():
            ws2.column_dimensions[cl].width = w
        ws2.freeze_panes = 'A3'
    else:
        ws2['A1'] = '⚠️ No se encontraron registros Non-Inventory Transfer'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
#  INTERFAZ STREAMLIT
# ══════════════════════════════════════════════════════════════════════════════

st.set_page_config(page_title='FLEXCON Extractor v3', page_icon='📦', layout='centered')

st.title('📦 FLEXCON PDF → Excel  (v3)')
st.markdown(
    'Sube un PDF **FLEXCON INV TRANSFER** (puede incluir páginas '
    '**Non-Inventory Transfer**). Se genera un Excel con dos hojas.'
)

uploaded_file = st.file_uploader('Selecciona el archivo PDF', type=['pdf'])

if uploaded_file:
    pdf_bytes = uploaded_file.read()

    with st.spinner('Extrayendo datos…'):
        inv_rec,    inv_meta    = extract_inventory_data(pdf_bytes)
        noninv_rec, noninv_meta = extract_non_inventory_data(pdf_bytes)

    c1, c2 = st.columns(2)
    with c1:
        st.subheader('🗂️ INV Transfer')
        st.metric('Registros', len(inv_rec))
        if inv_rec:
            df_inv = pd.DataFrame(inv_rec)
            df_inv['VALUE'] = df_inv['VALUE'].apply(safe_float)
            calc = df_inv['VALUE'].sum()
            st.metric('Valor calculado', f'${calc:,.2f}')
            decl = inv_meta.get('total_value') or 0
            if decl:
                diff = calc - decl
                st.metric('Total declarado', f'${decl:,.2f}',
                          delta='✅ OK' if abs(diff) < 1 else f'⚠️ Dif ${diff:+,.2f}')
    with c2:
        st.subheader('📋 Non-Inv Transfer')
        st.metric('Registros', len(noninv_rec))
        if noninv_rec:
            df_ni = pd.DataFrame(noninv_rec)
            df_ni['VALUE'] = df_ni['VALUE'].apply(safe_float)
            calc2 = df_ni['VALUE'].sum()
            st.metric('Valor calculado', f'${calc2:,.2f}')
            decl2 = noninv_meta.get('total_value') or 0
            if decl2:
                diff2 = calc2 - decl2
                st.metric('Total declarado', f'${decl2:,.2f}',
                          delta='✅ OK' if abs(diff2) < 1 else f'⚠️ Dif ${diff2:+,.2f}')

    if inv_rec:
        with st.expander('👁️ Preview — INV Transfer', expanded=False):
            st.dataframe(pd.DataFrame(inv_rec), use_container_width=True)
    if noninv_rec:
        with st.expander('👁️ Preview — Non-Inv Transfer', expanded=False):
            st.dataframe(pd.DataFrame(noninv_rec), use_container_width=True)

    if inv_rec or noninv_rec:
        with st.spinner('Generando Excel…'):
            xlsx = build_excel(inv_rec, inv_meta, noninv_rec, noninv_meta)
        st.success('✅ Excel listo')
        st.download_button(
            label='⬇️ Descargar Excel',
            data=xlsx,
            file_name=os.path.splitext(uploaded_file.name)[0] + '_EXTRACTED.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    else:
        st.error('❌ No se extrajeron datos. Verifica que sea un documento FLEXCON válido.')
else:
    st.info('👆 Sube un archivo PDF para comenzar.')
