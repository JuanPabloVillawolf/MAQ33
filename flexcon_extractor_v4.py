# -*- coding: utf-8 -*-
"""FLEXCON_Extractor_v5.py — Streamlit App
Cambios v5 vs v4:
  - Columnas simplificadas: solo Transfer Date, Page, Line, Item, Description,
    Origin, Quantity, UOM, QTY(M), Value, Doc Type
  - Bug fix: zona ITEM empieza en 69 (antes 72) → captura items a x0=71.7
  - Bug fix: ventana descripción Non-Inv ampliada (-15px a +25px) + límite inferior estricto
  - Bug fix: fix_value detecta coma decimal (40,00 → 40.00)
  - Bug fix: quantity solo acepta tokens numéricos puros
  - Bug fix: descripción Non-Inv no contamina fila siguiente
"""
import re, io, os, warnings
import pandas as pd
import pdfplumber
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")

# ── Zonas de columna ─────────────────────────────────────────────────────────
INV_COLS = {
    "LINE":(0,43),"B_P":(43,69),"ITEM":(69,143),"DESCRIPTION":(143,369),
    "ORIGIN":(369,427),"QUANTITY":(427,491),"UOM":(491,533),
    "QTY_M":(533,583),"VALUE":(712,9999),
}
NI_COLS = {
    "LINE":(0,100),"ITEM_NUMBER":(100,230),"DESCRIPTION":(230,483),
    "UOM":(483,515),"ORIGIN":(515,545),"QTY":(545,605),
    "LOT":(605,655),"VALUE":(655,710),
}
ORIGINS    = {"USA","JPN","NLD","MEX","CHN","KOR","TWN","DEU","PC"}
LOT_RE     = re.compile(r'^[A-Za-z][Oo0-9]{7,}$')
FORM_WORDS = {"shipped","received","shipping","notes","date","total",
              "value","boxes","by","transfer","truck","flexcon","seip"}

# Columnas de salida (simplificadas)
COLUMNS = ['Transfer Date','Page','Line','Item','Description',
           'Origin','Quantity','UOM','QTY(M)','Value','Doc Type']
COL_WIDTHS = {'Transfer Date':13,'Page':6,'Line':6,'Item':14,
              'Description':52,'Origin':8,'Quantity':11,'UOM':7,
              'QTY(M)':9,'Value':12,'Doc Type':22}

# ── Utilidades ───────────────────────────────────────────────────────────────
def col_of(x, zones):
    for k,(lo,hi) in zones.items():
        if lo<=x<hi: return k
    return None

def is_garbage(t):
    t=t.strip()
    if not t: return True
    return sum(1 for c in t if c.isalnum() or c in '$.,"-/()') / max(len(t),1) < 0.4

def fix_value(s):
    """
    Normaliza valores monetarios con OCR:
      2.279.90 → 2279.90   (doble punto = separador miles)
      3,582.70 → 3582.70   (coma = separador miles)
      40,00    → 40.00     (coma decimal: coma + exactamente 2 dígitos al final)
    """
    s = s.strip().replace('$','').replace(' ','').replace('—','')
    if not s: return None
    # Caso: dos o más puntos → todos menos el último son miles
    parts = s.split('.')
    if len(parts) >= 3:
        s = ''.join(parts[:-1]) + '.' + parts[-1]
    # Coma decimal: termina en coma + exactamente 2 dígitos sin punto previo
    m = re.match(r'^(\d[\d.]*)(?:,(\d{2}))$', s)
    if m:
        s = m.group(1).replace(',','') + '.' + m.group(2)
    else:
        # Coma como miles
        s = s.replace(',','')
    try:
        v = float(s)
        return v if v > 0 else None
    except: return None

def parse_qty(s):
    """Solo acepta tokens puramente numéricos."""
    s = s.strip().replace(',','')
    if re.match(r'^\d+$', s): return s
    # Punto como miles: e.g. "1.000" → 1000
    if re.match(r'^\d+\.\d{3,}$', s): return s.replace('.','')
    return ''

def ro(items, bucket=3):
    return [t for _,_,t in sorted(items, key=lambda w:(round(w[0]/bucket),w[1]))]

def ptext(page):
    try: return page.extract_text(x_tolerance=3,y_tolerance=3) or ""
    except: return page.extract_text() or ""

def pwords(page):
    try: return page.extract_words(x_tolerance=3,y_tolerance=3)
    except: return page.extract_words()

def get_clean_words(page, min_top=70):
    return [{'x0':w['x0'],'top':w['top'],'text':w['text'].strip()}
            for w in pwords(page)
            if w['text'].strip() and not is_garbage(w['text']) and w['top']>min_top]

# ── Detección de metadatos de página ─────────────────────────────────────────
def detect_meta(page):
    txt   = ptext(page)
    words = pwords(page)
    hw    = ' '.join(w['text'] for w in words if w['top']<55).upper()
    is_ni = any(m in txt.lower() for m in
                ['non-inventory transfer','non-lnventory','requestor name',
                 'seip - flexcon','seip -flexcon'])
    doc_type = 'NON_INV' if is_ni else 'INV'

    date = ''
    for pat in [r'Transfer Date[:\s]+(\d{1,2}/\d{1,2}/\d{2,4})',
                r'Shipment Dal[e;]+[:\s]*(\d{2}/\d{2}/\d{2,4})',
                r'Shipment Date[:\s]*(\d{2}/\d{2}/\d{2,4})']:
        m = re.search(pat, txt, re.I)
        if m: date = m.group(1).strip(); break
    if not date:
        m = re.search(r'\b(\d{1,2}/\d{1,2}/\d{2,4})\b', hw)
        if m: date = m.group(1)

    pg = 1
    m = re.search(r'Page[;:\s]+(\d+)', txt, re.I)
    if m: pg = int(m.group(1))

    ts = ''
    m = re.search(r'\b(\d{2}:\d{2}:\d{2})\b', txt)
    if m: ts = m.group(1)

    doc_id = ''
    m = re.search(r'\b(P\d{7,})\b', txt)
    if m: doc_id = m.group(1)

    tv = None
    for pat in [r'Total\s+Value[:\s]+\$?\s*([\d,.]+)',
                r'TOTAL\s+VALUE[;:\s]+\$?([\d,.]+)']:
        m = re.search(pat, txt, re.I)
        if m:
            tv = fix_value(m.group(1))
            if tv: break

    return dict(doc_type=doc_type,date=date,page_num=pg,
                timestamp=ts,doc_id=doc_id,total_val=tv)

# ── Extracción INV (una página) ───────────────────────────────────────────────
def extract_inv_page(page, meta):
    cw_list = get_clean_words(page, min_top=70)

    anchors = []
    for cw in cw_list:
        if col_of(cw['x0'], INV_COLS)=='LINE':
            if re.match(r'^\d{1,3}$', cw['text']) and 1<=int(cw['text'])<=99:
                anchors.append((cw['top'], int(cw['text'])))
    anchors.sort(key=lambda x:x[0])

    records = []
    for idx, (anchor_top, line_num) in enumerate(anchors):
        # Límite inferior estricto = siguiente ancla - 5px
        y_lo = anchor_top - 8
        y_hi = anchors[idx+1][0] - 5 if idx+1<len(anchors) else anchor_top + 30

        def coll(keys, dy_lo, dy_hi):
            r={k:[] for k in keys}
            for cw in cw_list:
                dy = cw['top'] - anchor_top
                if dy_lo <= dy <= dy_hi:
                    c=col_of(cw['x0'],INV_COLS)
                    if c in keys: r[c].append((cw['top'],cw['x0'],cw['text']))
            return r

        # Campos de la misma fila
        same   = coll(['B_P','ITEM','DESCRIPTION','ORIGIN','QUANTITY'], -8, 8)
        # UOM y QTY_M pueden flotar ±18px
        flt    = coll(['UOM','QTY_M','VALUE'], -6, 18)
        # LOT en sub-fila (+5 a +20px)
        below  = coll(['DESCRIPTION'], 5, 20)

        item = ' '.join(ro(same['ITEM']))
        desc_words = ro(same['DESCRIPTION'])
        # Añadir LOT de sub-fila si existe
        lot_words = [t for _,_,t in sorted(below['DESCRIPTION'],
                     key=lambda w:(round(w[0]/3),w[1]))
                     if LOT_RE.match(t.replace('O','0').replace('o','0'))]
        if lot_words:
            desc_words.append(lot_words[0])
        desc = ' '.join(desc_words)

        origin = next((t for _,_,t in sorted(same['ORIGIN'],
                       key=lambda w:(round(w[0]/3),w[1])) if t.upper() in ORIGINS),'')

        qlist  = [parse_qty(t) for _,_,t in sorted(same['QUANTITY'],
                  key=lambda w:(round(w[0]/3),w[1])) if parse_qty(t)]
        qty    = int(qlist[0]) if qlist else None

        ulist  = [t.upper().lstrip('-') for _,_,t in sorted(flt['UOM'],
                  key=lambda w:(round(w[0]/3),w[1]))
                  if not is_garbage(t) and len(t.strip('- '))<=4 and t.strip('- ').isalpha()]
        uom    = ulist[0] if ulist else ''

        qmlist = [parse_qty(t) for _,_,t in sorted(flt['QTY_M'],
                  key=lambda w:(round(w[0]/3),w[1])) if parse_qty(t)]
        qtym   = int(qmlist[0]) if qmlist else None

        vlist  = []
        for _,_,t in sorted(flt['VALUE'], key=lambda w:(round(w[0]/3),w[1])):
            v=fix_value(t)
            if v: vlist.append(v)
        value = vlist[-1] if vlist else None

        records.append({
            'Transfer Date': meta['date'],
            'Page':          meta['page_num'],
            'Line':          line_num,
            'Item':          item,
            'Description':   desc,
            'Origin':        origin.upper() if origin else '',
            'Quantity':      qty,
            'UOM':           uom,
            'QTY(M)':        qtym,
            'Value':         value,
            'Doc Type':      'INV Transfer',
            '_doc_seq':      meta.get('doc_seq',1),
            '_ts':           meta.get('timestamp',''),
        })
    return records

# ── Extracción Non-Inv (una página) ──────────────────────────────────────────
def extract_ni_page(page, meta):
    cw_list = get_clean_words(page, min_top=95)

    anchors = []
    for cw in cw_list:
        if col_of(cw['x0'],NI_COLS)=='LINE':
            if re.match(r'^\d{1,2}$',cw['text']) and 1<=int(cw['text'])<=20:
                anchors.append((cw['top'],int(cw['text'])))
    anchors.sort(key=lambda x:x[0])

    records = []
    for i,(anchor_top,line_num) in enumerate(anchors):
        # Límite inferior estricto = siguiente ancla - 3px
        y_hi = anchors[i+1][0] - 3 if i+1<len(anchors) else anchor_top + 30
        # Ventana ESTRECHA para campos fijos
        lo_strict = anchor_top - 5
        # Ventana AMPLIA para descripción: máx entre 15px arriba del ancla
        # y 2px después del y_hi de la fila anterior (evita bleeding)
        prev_y_hi = anchors[i-1][0] + (anchor_top - anchors[i-1][0]) * 0.4 if i>0 else anchor_top - 20
        lo_wide   = max(anchor_top - 15, prev_y_hi - 1)

        def coll_range(keys, lo, hi):
            r={}
            for cw in cw_list:
                if lo<=cw['top']<=hi:
                    c=col_of(cw['x0'],NI_COLS)
                    if c in keys:
                        r.setdefault(c,[]).append((cw['top'],cw['x0'],cw['text']))
            return r

        strict = coll_range(['ITEM_NUMBER','UOM','ORIGIN','QTY','LOT','VALUE'], lo_strict, y_hi)
        wide   = coll_range(['DESCRIPTION'], lo_wide, y_hi)

        item = ' '.join(ro(strict.get('ITEM_NUMBER',[])))
        if item.upper() in ('N/A','NA'): item = 'N/A'

        desc = ' '.join(ro(wide.get('DESCRIPTION',[])))
        uom  = ' '.join(t for _,_,t in sorted(strict.get('UOM',[]),
               key=lambda w:(round(w[0]/3),w[1])) if t.isalpha() and len(t)<=4).upper()

        origin = next((t for _,_,t in strict.get('ORIGIN',[])
                       if t.upper() in ORIGINS - {'PC'}),'')

        qlist = [parse_qty(t) for _,_,t in strict.get('QTY',[])
                 if parse_qty(t) and int(parse_qty(t)) > 0]
        qty   = int(qlist[0]) if qlist else None

        lot   = next((t for _,_,t in strict.get('LOT',[])
                      if t.upper() in ('N/A','NA')),'N/A')

        vlist = []
        for _,_,t in strict.get('VALUE',[]):
            v=fix_value(t.replace('$',''))
            if v: vlist.append(v)
        value = vlist[0] if vlist else None

        # Saltar líneas vacías
        if not item.strip() and not desc.strip(): continue
        if any(fw in item.lower() for fw in FORM_WORDS): continue

        records.append({
            'Transfer Date': meta['date'],
            'Page':          meta['page_num'],
            'Line':          line_num,
            'Item':          item,
            'Description':   desc,
            'Origin':        origin.upper() if origin else '',
            'Quantity':      qty,
            'UOM':           uom,
            'QTY(M)':        None,
            'Value':         value,
            'Doc Type':      'Non-Inventory Transfer',
            '_doc_seq':      meta.get('doc_seq',1),
            '_ts':           meta.get('timestamp',''),
        })
    return records

# ── Pipeline principal ────────────────────────────────────────────────────────
def extract_all(pdf_bytes):
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        page_metas = [(p, detect_meta(p)) for p in pdf.pages]

    cur_ts, cur_seq = None, 0
    for _, meta in page_metas:
        if meta['timestamp'] != cur_ts:
            cur_seq += 1; cur_ts = meta['timestamp']
        meta['doc_seq'] = cur_seq

    all_records   = []
    doc_summaries = {}

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page_obj, meta in zip(pdf.pages, [m for _,m in page_metas]):
            recs = extract_ni_page(page_obj, meta) if meta['doc_type']=='NON_INV' \
                   else extract_inv_page(page_obj, meta)
            all_records.extend(recs)
            seq = meta['doc_seq']
            if seq not in doc_summaries:
                doc_summaries[seq] = dict(doc_seq=seq, doc_id=meta['doc_id'],
                    date=meta['date'], doc_type=meta['doc_type'],
                    timestamp=meta['timestamp'], total_val=None)
            if meta['total_val']:
                doc_summaries[seq]['total_val'] = meta['total_val']

    return all_records, list(doc_summaries.values())

# ── Excel ─────────────────────────────────────────────────────────────────────
def _fill(h): return PatternFill(start_color=h,end_color=h,fill_type='solid')
def _side(s='thin',c='C0C0C0'): return Side(style=s,color=c)

def write_sheet(ws, df, header_hex, alt_hex, total_hex, title):
    thick = _side('medium','404040')
    ws.append([title])
    ws.merge_cells(f'A1:{get_column_letter(len(COLUMNS))}1')
    ws['A1'].font=Font(bold=True,italic=True,name='Arial',size=10,color='404040')
    ws['A1'].alignment=Alignment(horizontal='left',vertical='center')
    ws.row_dimensions[1].height=18
    ws.append(COLUMNS)
    for cell in ws[2]:
        cell.fill=_fill(header_hex)
        cell.font=Font(bold=True,color='FFFFFF',name='Arial',size=10)
        cell.border=Border(left=thick,right=thick,top=thick,bottom=thick)
        cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    ws.row_dimensions[2].height=28
    alt=_fill(alt_hex)
    bord=Border(left=_side(),right=_side(),top=_side(),bottom=_side())
    for ri,(_,row) in enumerate(df.iterrows(),3):
        ws.append([row.get(c,'') if pd.notna(row.get(c,'')) else '' for c in COLUMNS])
        for cell in ws[ri]:
            cell.border=bord; cell.font=Font(name='Arial',size=9)
            cell.alignment=Alignment(vertical='center')
            if ri%2==0: cell.fill=alt
    hdr={c.value:c.column_letter for c in ws[2]}
    ds,de=3,2+len(df)
    for ri in range(ds,de+1):
        for col in ['Quantity','QTY(M)']:
            if col in hdr: ws[f"{hdr[col]}{ri}"].number_format='#,##0'
        if 'Value' in hdr: ws[f"{hdr['Value']}{ri}"].number_format='$#,##0.00'
    tr=de+1
    ws.append(['TOTALES']+['']*( len(COLUMNS)-1))
    for col in ['Quantity','QTY(M)']:
        if col in hdr:
            cl=hdr[col]; ws[f'{cl}{tr}']=f'=IFERROR(SUM({cl}{ds}:{cl}{de}),"")'; ws[f'{cl}{tr}'].number_format='#,##0'
    if 'Value' in hdr:
        cl=hdr['Value']; ws[f'{cl}{tr}']=f'=IFERROR(SUM({cl}{ds}:{cl}{de}),"")'; ws[f'{cl}{tr}'].number_format='$#,##0.00'
    for cell in ws[tr]:
        cell.fill=_fill(total_hex); cell.font=Font(bold=True,name='Arial',size=10)
        cell.border=Border(left=thick,right=thick,top=thick,bottom=thick)
    for i,col in enumerate(COLUMNS,1):
        ws.column_dimensions[get_column_letter(i)].width=COL_WIDTHS.get(col,12)
    ws.freeze_panes='A3'
    ws.auto_filter.ref=f'A2:{get_column_letter(len(COLUMNS))}{de}'

def build_excel(all_records, doc_summaries):
    wb=Workbook()
    df_all=pd.DataFrame(all_records)
    for c in COLUMNS:
        if c not in df_all.columns: df_all[c]=''

    ws1=wb.active; ws1.title='Todos los Registros'
    doc_info='  |  '.join(
        f"Doc{d['doc_seq']}: {d['date']} {d['doc_type']}"
        +(f" TV=${d['total_val']:,.2f}" if d['total_val'] else '')
        for d in doc_summaries)
    write_sheet(ws1,df_all,'1F4E78','EEF3F8','C9D8EC',
                f'FLEXCON — Todos los registros en orden físico  |  {doc_info}')

    colors=[('2E5FA3','EEF3F8','C9D8EC'),('1A6B3C','EDF7EF','C6E8CC'),
            ('7B3F9E','F5EEF8','DCC6EC'),('B05000','FEF3E8','F5D5B0')]
    for dm in doc_summaries:
        seq=dm['doc_seq']
        df_d=df_all[df_all['_doc_seq']==seq].copy() if '_doc_seq' in df_all.columns \
             else df_all[df_all['Doc Type']==('INV Transfer' if dm['doc_type']=='INV' else 'Non-Inventory Transfer')].copy()
        if df_d.empty: continue
        dt=dm['date'].replace('/','_')
        dtype='INV' if dm['doc_type']=='INV' else 'NonInv'
        name=f"Doc{seq} {dt} {dtype}"[:31]
        ws=wb.create_sheet(name)
        hc,ac,tc=colors[(seq-1)%len(colors)]
        tv=f'  Total: ${dm["total_val"]:,.2f}' if dm['total_val'] else ''
        write_sheet(ws,df_d,hc,ac,tc,
                    f'{dm["doc_id"]}  {dm["date"]}  {dm["doc_type"]}{tv}')

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()

# ── Streamlit UI ──────────────────────────────────────────────────────────────
st.set_page_config(page_title='FLEXCON Extractor v5',page_icon='📦',layout='wide')
st.title('📦 FLEXCON PDF → Excel  (v5)')
st.markdown('Extrae en **orden físico exacto**. Detecta múltiples documentos por PDF.')

uploaded_file=st.file_uploader('Selecciona el archivo PDF',type=['pdf'])

if uploaded_file:
    pdf_bytes=uploaded_file.read()
    with st.spinner('Procesando…'):
        all_records,doc_summaries=extract_all(pdf_bytes)

    if not all_records:
        st.error('❌ No se extrajeron registros.')
    else:
        df=pd.DataFrame(all_records)
        st.subheader(f'📋 {len(doc_summaries)} documento(s)  |  {len(df)} registros')
        cols_ui=st.columns(min(len(doc_summaries),4))
        for i,dm in enumerate(doc_summaries):
            key = '_doc_seq' if '_doc_seq' in df.columns else 'Doc Type'
            df_d = df[df['_doc_seq']==dm['doc_seq']] if '_doc_seq' in df.columns else df
            calc=df_d['Value'].apply(lambda v: v if isinstance(v,(int,float)) else 0).sum()
            with cols_ui[i%len(cols_ui)]:
                icon='🗂️' if dm['doc_type']=='INV' else '📋'
                st.metric(f"{icon} Doc {dm['doc_seq']} — {dm['date']}",
                          f"{len(df_d)} registros", delta=f"${calc:,.2f}")
                if dm['total_val']:
                    diff=calc-dm['total_val']
                    st.caption(f"Decl: ${dm['total_val']:,.2f} — "
                               f"{'✅' if abs(diff)<1 else f'⚠️ Dif ${diff:+,.2f}'}")

        with st.expander('👁️ Preview',expanded=True):
            show=[c for c in COLUMNS if c in df.columns]
            st.dataframe(df[show],use_container_width=True,height=400)

        with st.spinner('Generando Excel…'):
            xlsx=build_excel(all_records,doc_summaries)
        st.success(f'✅ {len(all_records)} registros listos')
        st.download_button('⬇️ Descargar Excel',data=xlsx,
            file_name=os.path.splitext(uploaded_file.name)[0]+'_v5.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
else:
    st.info('👆 Sube un archivo PDF para comenzar.')
