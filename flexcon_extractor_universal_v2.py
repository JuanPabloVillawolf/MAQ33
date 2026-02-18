# -*- coding: utf-8 -*-
"""
FLEXCON_Extractor_Universal_v2.py
==================================
Extractor mejorado para documentos FLEXCON PDF → Excel

Mejoras respecto a v1:
  ✅ Detección de tipo por PÁGINA (no solo por documento)
  ✅ Manejo de documentos mixtos (Inventory + Non-Inventory en el mismo PDF)
  ✅ Hoja 1: FLEXCON INV TRANSFER
  ✅ Hoja 2: SEIP Non-Inventory Transfer
  ✅ Mejor regex para números de LOT (W01997391006, P00048352018, etc.)
  ✅ Reconstrucción de líneas partidas por pdfplumber
  ✅ BOXES / WEIGHT: marcados como 'N/D' cuando son manuscritos e ilegibles
  ✅ Extracción de metadata del documento (No. doc, fecha, total)
  ✅ Fila de TOTALES automática en cada hoja
  ✅ Formato profesional en ambas hojas
"""

# ─── Instalación de dependencias ─────────────────────────────────────────────
import subprocess, sys

def pip_install(pkg):
    subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])

for pkg in ["pandas", "openpyxl", "pdfplumber"]:
    pip_install(pkg)

print("✅ Librerías listas")

# ─── Importaciones ────────────────────────────────────────────────────────────
import re
import os
import warnings
import pandas as pd
import pdfplumber
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ─── Carga del archivo ────────────────────────────────────────────────────────
try:
    from google.colab import files
    print("\n📤 Sube tu archivo PDF de FLEXCON:")
    uploaded = files.upload()
    PDF_PATH = list(uploaded.keys())[0]
except ImportError:
    # Modo local: pasar ruta como argumento o variable
    import sys
    PDF_PATH = sys.argv[1] if len(sys.argv) > 1 else "flexcon.pdf"

print(f"✅ Procesando: {PDF_PATH}")

# ─── CONSTANTES ───────────────────────────────────────────────────────────────
ORIGINS       = {"USA", "JPN", "NLD", "MEX", "CHN", "KOR", "TWN", "DEU"}
UOM_TYPES     = {"PC", "FT", "EA", "GAL", "PT", "IN", "BOX", "KG", "LB", "M", "SET"}
LOT_PATTERN   = re.compile(r"^[A-Z]\d{7,}", re.IGNORECASE)   # P00048352018, W01997391006…
VALUE_PATTERN = re.compile(r"\$?([\d,]+\.\d{2})$")


# ══════════════════════════════════════════════════════════════════════════════
#  UTILIDADES COMUNES
# ══════════════════════════════════════════════════════════════════════════════

def clean_num(s: str) -> str:
    """Quita comas y símbolo $ de un string numérico."""
    return s.replace(",", "").replace("$", "").strip()

def safe_int(val):
    try:
        if pd.isna(val) or str(val).strip() in ("", "N/A", "N/D"):
            return None
        return int(clean_num(str(val)))
    except:
        return None

def safe_float(val):
    try:
        if pd.isna(val) or str(val).strip() in ("", "N/A", "N/D"):
            return None
        return float(clean_num(str(val)))
    except:
        return None

def extract_page_text(page) -> str:
    """Extrae texto de la página con manejo de errores."""
    try:
        return page.extract_text(x_tolerance=3, y_tolerance=3) or ""
    except:
        return page.extract_text() or ""

def is_non_inventory_page(text: str) -> bool:
    """Detecta si una página es del formato Non-Inventory Transfer."""
    markers = [
        "Non-Inventory Transfer",
        "non-inventory",
        "NON-INVENTORY",
        "Requestor Name",
        "REQUESTOR",
        "Ship To Address",
        "M27009",           # número de formulario Non-Inv
    ]
    return any(m.lower() in text.lower() for m in markers)

def is_inventory_page(text: str) -> bool:
    """Detecta si una página es del formato INV TRANSFER estándar."""
    markers = ["INV TRANSFER", "INVENTORY TRANSFER", "FLEXCON INV", "P5546"]
    return any(m.upper() in text.upper() for m in markers)

def extract_document_meta(text: str) -> dict:
    """Extrae número de documento, fecha y totales del texto de la página."""
    meta = {"doc_number": "", "transfer_date": "", "total_value": None, "total_boxes": None}

    # Número de documento (P5546038, P55…)
    m = re.search(r"\b(P\d{7,})\b", text)
    if m:
        meta["doc_number"] = m.group(1)

    # Fecha de transferencia
    m = re.search(r"Transfer Date[:\s]+(\d{1,2}/\d{1,2}/\d{2,4})", text, re.IGNORECASE)
    if m:
        meta["transfer_date"] = m.group(1)
    else:
        m = re.search(r"\b(\d{1,2}/\d{1,2}/\d{2,4})\b", text)
        if m:
            meta["transfer_date"] = m.group(1)

    # Total Value
    m = re.search(r"Total\s+Value[:\s]+\$?([\d,]+\.\d{2})", text, re.IGNORECASE)
    if m:
        meta["total_value"] = float(clean_num(m.group(1)))

    # Total Boxes
    m = re.search(r"Total\s+Boxes[:\s]+(\d+)", text, re.IGNORECASE)
    if m:
        meta["total_boxes"] = int(m.group(1))

    return meta


# ══════════════════════════════════════════════════════════════════════════════
#  EXTRACCIÓN — INVENTORY TRANSFER (páginas 1-N)
# ══════════════════════════════════════════════════════════════════════════════

def parse_inventory_line(line_parts: list, line_num: str, bp: str, item: str) -> dict | None:
    """
    Dado un bloque de tokens que sigue a LINE  B/P  ITEM, extrae:
    DESCRIPTION, ORIGIN, QUANTITY, UOM, QTY(M), VALUE
    BOXES y WEIGHT son generalmente manuscritos → se marcan N/D
    """
    # Encontrar posición del ORIGIN dentro de los tokens
    origin_idx = -1
    for i, tok in enumerate(line_parts):
        if tok.upper() in ORIGINS:
            origin_idx = i
            break

    if origin_idx == -1:
        return None

    description = " ".join(line_parts[:origin_idx])
    origin      = line_parts[origin_idx].upper()
    after       = [clean_num(t) for t in line_parts[origin_idx + 1:]]

    # QUANTITY (primer token después de ORIGIN que sea entero)
    quantity, uom, qty_m, value = "", "", "", ""
    numeric_after = []
    for t in after:
        if re.match(r"^\d+$", t):
            numeric_after.append(("int", t))
        elif re.match(r"^\d+\.\d{2}$", t):
            numeric_after.append(("float", t))
        elif t.upper() in UOM_TYPES:
            numeric_after.append(("uom", t))
        else:
            numeric_after.append(("other", t))

    # Reconstruir campos en orden esperado: QTY  UOM  QTY(M)  [boxes hw]  [weight hw]  VALUE
    int_vals   = [v for k, v in numeric_after if k == "int"]
    float_vals = [v for k, v in numeric_after if k == "float"]
    uom_vals   = [v for k, v in numeric_after if k == "uom"]

    if int_vals:
        quantity = int_vals[0]
    if uom_vals:
        uom = uom_vals[0].upper()
    if len(int_vals) >= 2:
        qty_m = int_vals[1]           # segundo entero = QTY(M)
    if float_vals:
        value = float_vals[-1]        # último decimal = VALUE

    return {
        "LINE":           line_num,
        "B/P":            bp,
        "ITEM":           item,
        "DESCRIPTION/LOT": description.strip(),
        "ORIGIN":         origin,
        "QUANTITY":       quantity,
        "UOM":            uom,
        "QTY(M)":         qty_m,
        "BOXES":          "N/D",     # manuscrito
        "WEIGHT(KG)":     "N/D",     # manuscrito
        "VALUE":          value,
    }


def extract_inventory_data(pdf_path: str) -> tuple[list, dict]:
    """
    Extrae registros de las páginas de tipo INV TRANSFER.
    Retorna (lista_de_registros, metadata_del_documento).
    """
    all_records = []
    doc_meta    = {}

    # Encabezado de línea de datos: empieza con 1-3 dígitos, luego 3 dígitos (B/P), luego item
    DATA_LINE_RE = re.compile(r"^(\d{1,3})\s+(\d{3})\s+([\w\-\.]+)(.*)")

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            text = extract_page_text(page)
            if not text:
                continue
            if is_non_inventory_page(text):
                continue    # Esa página se procesa aparte
            if not is_inventory_page(text) and page_num > 1:
                # Páginas intermedias sin header: asumir que son continuación
                pass

            # Extraer metadata solo de la primera página de inventario
            if not doc_meta:
                doc_meta = extract_document_meta(text)

            lines = text.split("\n")

            # Construir lista de líneas "limpias" y detectar continuaciones
            i = 0
            while i < len(lines):
                raw = lines[i].strip()
                m = DATA_LINE_RE.match(raw)
                if not m:
                    i += 1
                    continue

                line_num = m.group(1)
                bp       = m.group(2)
                item     = m.group(3)
                rest     = m.group(4).strip()

                # Acumular tokens: puede haber LOT y/o datos en líneas siguientes
                tokens = rest.split()
                j = i + 1

                # Ventana de lookahead: máx 4 líneas para recolectar datos
                while j < min(i + 5, len(lines)):
                    candidate = lines[j].strip()

                    # Es un LOT si coincide con el patrón
                    if LOT_PATTERN.match(candidate) and not any(
                        candidate.upper().startswith(orig) for orig in ORIGINS
                    ):
                        # Agregar LOT a la descripción
                        tokens = [candidate] + tokens if not tokens else tokens + [candidate]
                        j += 1
                        continue

                    # Si la línea contiene un ORIGIN, son los datos numéricos
                    if any(f" {o} " in f" {candidate} " or candidate.startswith(o + " ")
                           for o in ORIGINS):
                        tokens += candidate.split()
                        j += 1
                        break

                    # Si la línea empieza con dígitos (puede ser qty u otra línea de datos)
                    if re.match(r"^\d", candidate):
                        # Si ya tenemos ORIGIN en los tokens, parar
                        if any(t.upper() in ORIGINS for t in tokens):
                            break
                        tokens += candidate.split()
                        j += 1
                        continue

                    break

                record = parse_inventory_line(tokens, line_num, bp, item)
                if record:
                    all_records.append(record)

                i = j   # saltar líneas ya procesadas

    return all_records, doc_meta


# ══════════════════════════════════════════════════════════════════════════════
#  EXTRACCIÓN — NON-INVENTORY TRANSFER (última página / páginas con ese header)
# ══════════════════════════════════════════════════════════════════════════════

def extract_non_inventory_data(pdf_path: str) -> tuple[list, dict]:
    """
    Extrae registros de páginas del tipo SEIP Non-Inventory Transfer.
    Columnas: Line, Item Number, Description, U/M, Orig, Qty, Lot(N/A), Value, Boxes, Weight
    """
    all_records = []
    doc_meta    = {}

    # Patrón de datos: número de línea + item + descripción + UOM + ORIGIN + qty + N/A + valor
    # Formato típico: "1  8289R  Visi Mark - For crimp striper Red  PT  USA  4  N/A  $466.12"
    # Los campos BOXES y WEIGHT son manuscritos al final → ignorados

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            text = extract_page_text(page)
            if not text or not is_non_inventory_page(text):
                continue

            if not doc_meta:
                # Meta del Non-Inv doc
                doc_meta["page"] = page_num
                m = re.search(r"Requestor Name[:\s]+(.+)", text, re.IGNORECASE)
                if m:
                    doc_meta["requestor"] = m.group(1).strip()
                m = re.search(r"From Location[:\s]+(.+)", text, re.IGNORECASE)
                if m:
                    doc_meta["from_location"] = m.group(1).strip()
                m = re.search(r"Shipment Date[:\s]+(\d{1,2}/\d{1,2}/\d{2,4})", text, re.IGNORECASE)
                if m:
                    doc_meta["shipment_date"] = m.group(1)
                m = re.search(r"Ship To Address[:\s]+(.+)", text, re.IGNORECASE)
                if m:
                    doc_meta["ship_to"] = m.group(1).strip()
                m = re.search(r"TOTAL\s+VALUE[:\s]+\$?([\d,]+\.\d{2})", text, re.IGNORECASE)
                if m:
                    doc_meta["total_value"] = float(clean_num(m.group(1)))

            lines = text.split("\n")

            for line in lines:
                # La línea debe empezar con 1-2 dígitos (número de línea 1-17)
                m = re.match(r"^(\d{1,2})\s+(\S+)\s+(.+)", line.strip())
                if not m:
                    continue

                line_num = int(m.group(1))
                if line_num < 1 or line_num > 20:
                    continue

                item_num    = m.group(2)
                remainder   = m.group(3).strip()

                # Buscar ORIGIN de derecha a izquierda para separar descripción del resto
                origin_match = None
                for orig in ORIGINS:
                    # Buscamos el último ORIGIN en la línea
                    pat = re.search(rf"\b{orig}\b(.*)", remainder, re.IGNORECASE)
                    if pat:
                        origin_match = (orig.upper(), pat.start(), pat.group(1).strip())
                        break

                if not origin_match:
                    continue

                origin          = origin_match[0]
                before_origin   = remainder[:origin_match[1]].strip()
                after_origin    = origin_match[2]

                # Antes del ORIGIN: "DESCRIPCIÓN  UOM"
                # El UOM está justo antes del ORIGIN
                uom = ""
                description = before_origin
                words_before = before_origin.split()
                if words_before and words_before[-1].upper() in UOM_TYPES:
                    uom         = words_before[-1].upper()
                    description = " ".join(words_before[:-1])

                # Después del ORIGIN: "QTY  N/A  $VALUE"
                after_tokens = after_origin.split()
                after_tokens = [clean_num(t) for t in after_tokens]

                quantity = ""
                lot      = "N/A"
                value    = ""

                # Filtrar tokens
                for tok in after_tokens:
                    if tok.upper() in ("N/A", "NA"):
                        lot = "N/A"
                    elif re.match(r"^\d+\.\d{2}$", tok):
                        value = tok
                    elif re.match(r"^\d+$", tok) and not quantity:
                        quantity = tok

                record = {
                    "LINE":         str(line_num),
                    "ITEM NUMBER":  item_num,
                    "DESCRIPTION":  description.strip(),
                    "U/M":          uom,
                    "ORIGIN":       origin,
                    "QTY REQUESTED": quantity,
                    "LOT NUMBER":   lot,
                    "VALUE":        value,
                    "BOXES":        "N/D",   # manuscrito
                    "WEIGHT(KG)":   "N/D",   # manuscrito
                }
                all_records.append(record)

    return all_records, doc_meta


# ══════════════════════════════════════════════════════════════════════════════
#  CONSTRUCCIÓN DEL EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def apply_sheet_style(ws, df: pd.DataFrame,
                      header_color: str,
                      col_widths: dict,
                      money_col: str,
                      qty_cols: list[str]):
    """Aplica formato profesional a una hoja de Excel."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    thin = Side(style="thin", color="C0C0C0")
    thick = Side(style="medium", color="404040")
    thin_border  = Border(left=thin, right=thin, top=thin, bottom=thin)
    thick_border = Border(left=thick, right=thick, top=thick, bottom=thick)

    h_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    h_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    t_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")  # total row
    alt_fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")

    # Header row
    for cell in ws[1]:
        cell.fill      = h_fill
        cell.font      = h_font
        cell.border    = thick_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30

    # Obtener letra de cada columna por nombre
    headers = {cell.value: cell.column_letter for cell in ws[1]}

    # Filas de datos
    for row_idx in range(2, len(df) + 2):
        is_alt = (row_idx % 2 == 0)
        for cell in ws[row_idx]:
            cell.border    = thin_border
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if is_alt:
                cell.fill = alt_fill

    # Fila de total (última fila después de los datos)
    total_row = len(df) + 2
    ws.append([""] * len(df.columns))   # fila vacía como separador
    total_row_actual = ws.max_row
    for cell in ws[total_row_actual]:
        cell.fill   = t_fill
        cell.font   = Font(bold=True, name="Arial", size=9)
        cell.border = thick_border

    # Escritura de totales con fórmulas Excel
    for col_name in qty_cols:
        if col_name in headers:
            col_l = headers[col_name]
            ws[f"{col_l}{total_row_actual}"] = \
                f'=IFERROR(SUM({col_l}2:{col_l}{total_row-1}),"")'

    if money_col in headers:
        col_l = headers[money_col]
        ws[f"{col_l}{total_row_actual}"] = \
            f'=IFERROR(SUM({col_l}2:{col_l}{total_row-1}),"")'

    # Etiqueta TOTAL
    first_col = ws[f"A{total_row_actual}"]
    first_col.value = "TOTAL"
    first_col.font  = Font(bold=True, name="Arial", size=9)

    # Formato de moneda y números
    for row_idx in range(2, total_row_actual + 1):
        if money_col in headers:
            c = ws[f"{headers[money_col]}{row_idx}"]
            if c.value is not None:
                c.number_format = '$#,##0.00'
        for col_name in qty_cols:
            if col_name in headers:
                c = ws[f"{headers[col_name]}{row_idx}"]
                if c.value is not None:
                    c.number_format = '#,##0'

    # Ancho de columnas
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze header
    ws.freeze_panes = "A2"


def build_excel(inv_records: list, inv_meta: dict,
                noninv_records: list, noninv_meta: dict,
                output_path: str):
    """Construye el archivo Excel con dos hojas."""
    from openpyxl import Workbook

    wb = Workbook()

    # ── Hoja 1: Inventory Transfer ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "INV Transfer"

    if inv_records:
        df1 = pd.DataFrame(inv_records)

        # Convertir tipos
        df1["LINE"]     = df1["LINE"].apply(safe_int)
        df1["QUANTITY"] = df1["QUANTITY"].apply(safe_int)
        df1["QTY(M)"]   = df1["QTY(M)"].apply(safe_int)
        df1["VALUE"]    = df1["VALUE"].apply(safe_float)

        df1.sort_values("LINE", inplace=True, ignore_index=True)

        # Fila de metadata encima de los datos
        meta_str = (
            f"Doc: {inv_meta.get('doc_number','')}   "
            f"Fecha: {inv_meta.get('transfer_date','')}   "
            f"Total Declarado: ${inv_meta.get('total_value', 0):,.2f}"
        )
        ws1.append([meta_str])
        ws1.merge_cells(f"A1:{get_column_letter(len(df1.columns))}1")
        ws1["A1"].font = Font(bold=True, italic=True, name="Arial", size=9, color="404040")
        ws1["A1"].alignment = Alignment(horizontal="left")

        # Escribir encabezados en fila 2
        ws1.append(list(df1.columns))

        # Escribir datos desde fila 3
        for _, row in df1.iterrows():
            ws1.append(list(row))

        # Re-mapear estilos considerando que los encabezados están en fila 2
        # (hacemos override manual de la fila de encabezado)
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        thin   = Side(style="thin",   color="C0C0C0")
        thick  = Side(style="medium", color="404040")
        h_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        t_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        alt_fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")

        # Estilo de la fila de headers (fila 2)
        for cell in ws1[2]:
            cell.fill      = h_fill
            cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            cell.border    = Border(left=thick, right=thick, top=thick, bottom=thick)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[2].height = 28

        # Estilo de filas de datos (fila 3 en adelante)
        data_start = 3
        data_end   = data_start + len(df1) - 1
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for row_idx in range(data_start, data_end + 1):
            is_alt = (row_idx % 2 == 0)
            for cell in ws1[row_idx]:
                cell.border = thin_border
                cell.font   = Font(name="Arial", size=9)
                cell.alignment = Alignment(vertical="center")
                if is_alt:
                    cell.fill = alt_fill

        # Fila de totales
        total_row = data_end + 1
        ws1.append(["TOTAL"] + [""] * (len(df1.columns) - 1))
        headers_row = {cell.value: cell.column_letter for cell in ws1[2]}

        qty_cols_inv = ["QUANTITY", "QTY(M)"]
        for col_name in qty_cols_inv:
            if col_name in headers_row:
                col_l = headers_row[col_name]
                ws1[f"{col_l}{total_row}"] = \
                    f"=IFERROR(SUM({col_l}{data_start}:{col_l}{data_end}),\"\")"
                ws1[f"{col_l}{total_row}"].number_format = "#,##0"

        if "VALUE" in headers_row:
            col_l = headers_row["VALUE"]
            ws1[f"{col_l}{total_row}"] = \
                f"=IFERROR(SUM({col_l}{data_start}:{col_l}{data_end}),\"\")"
            ws1[f"{col_l}{total_row}"].number_format = "$#,##0.00"

        for cell in ws1[total_row]:
            cell.fill   = t_fill
            cell.font   = Font(bold=True, name="Arial", size=9)
            cell.border = Border(left=thick, right=thick, top=thick, bottom=thick)

        # Formato de moneda y número en datos
        for row_idx in range(data_start, data_end + 1):
            if "VALUE" in headers_row:
                ws1[f"{headers_row['VALUE']}{row_idx}"].number_format = "$#,##0.00"
            for c in qty_cols_inv:
                if c in headers_row:
                    ws1[f"{headers_row[c]}{row_idx}"].number_format = "#,##0"

        # Ancho de columnas
        col_widths_inv = {
            "A": 6,  "B": 7,  "C": 16, "D": 58,
            "E": 9,  "F": 12, "G": 7,  "H": 10,
            "I": 8,  "J": 11, "K": 13,
        }
        for col_l, width in col_widths_inv.items():
            ws1.column_dimensions[col_l].width = width

        ws1.freeze_panes = "A3"
    else:
        ws1["A1"] = "⚠️ No se encontraron registros de tipo INV TRANSFER"

    # ── Hoja 2: Non-Inventory Transfer ──────────────────────────────────────
    ws2 = wb.create_sheet("Non-Inv Transfer")

    if noninv_records:
        df2 = pd.DataFrame(noninv_records)
        df2["LINE"] = df2["LINE"].apply(safe_int)
        df2["QTY REQUESTED"] = df2["QTY REQUESTED"].apply(safe_int)
        df2["VALUE"] = df2["VALUE"].apply(safe_float)
        df2.sort_values("LINE", inplace=True, ignore_index=True)

        meta_str2 = (
            f"Requestor: {noninv_meta.get('requestor','')}   "
            f"From: {noninv_meta.get('from_location','')}   "
            f"Ship To: {noninv_meta.get('ship_to','')}   "
            f"Fecha: {noninv_meta.get('shipment_date','')}   "
            f"Total Declarado: ${noninv_meta.get('total_value', 0):,.2f}"
        )
        ws2.append([meta_str2])
        ws2.merge_cells(f"A1:{get_column_letter(len(df2.columns))}1")
        ws2["A1"].font = Font(bold=True, italic=True, name="Arial", size=9, color="404040")
        ws2["A1"].alignment = Alignment(horizontal="left")

        ws2.append(list(df2.columns))

        for _, row in df2.iterrows():
            ws2.append(list(row))

        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        thin   = Side(style="thin",   color="C0C0C0")
        thick  = Side(style="medium", color="404040")
        h_fill = PatternFill(start_color="1D6A3A", end_color="1D6A3A", fill_type="solid")  # verde
        t_fill = PatternFill(start_color="D5F0DC", end_color="D5F0DC", fill_type="solid")
        alt_fill = PatternFill(start_color="EDF7EF", end_color="EDF7EF", fill_type="solid")

        for cell in ws2[2]:
            cell.fill      = h_fill
            cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            cell.border    = Border(left=thick, right=thick, top=thick, bottom=thick)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[2].height = 28

        data_start2 = 3
        data_end2   = data_start2 + len(df2) - 1
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for row_idx in range(data_start2, data_end2 + 1):
            is_alt = (row_idx % 2 == 0)
            for cell in ws2[row_idx]:
                cell.border = thin_border
                cell.font   = Font(name="Arial", size=9)
                cell.alignment = Alignment(vertical="center")
                if is_alt:
                    cell.fill = alt_fill

        total_row2 = data_end2 + 1
        ws2.append(["TOTAL"] + [""] * (len(df2.columns) - 1))
        headers_row2 = {cell.value: cell.column_letter for cell in ws2[2]}

        if "QTY REQUESTED" in headers_row2:
            col_l = headers_row2["QTY REQUESTED"]
            ws2[f"{col_l}{total_row2}"] = \
                f"=IFERROR(SUM({col_l}{data_start2}:{col_l}{data_end2}),\"\")"
            ws2[f"{col_l}{total_row2}"].number_format = "#,##0"

        if "VALUE" in headers_row2:
            col_l = headers_row2["VALUE"]
            ws2[f"{col_l}{total_row2}"] = \
                f"=IFERROR(SUM({col_l}{data_start2}:{col_l}{data_end2}),\"\")"
            ws2[f"{col_l}{total_row2}"].number_format = "$#,##0.00"

        for cell in ws2[total_row2]:
            cell.fill   = t_fill
            cell.font   = Font(bold=True, name="Arial", size=9)
            cell.border = Border(left=thick, right=thick, top=thick, bottom=thick)

        for row_idx in range(data_start2, data_end2 + 1):
            if "VALUE" in headers_row2:
                ws2[f"{headers_row2['VALUE']}{row_idx}"].number_format = "$#,##0.00"
            if "QTY REQUESTED" in headers_row2:
                ws2[f"{headers_row2['QTY REQUESTED']}{row_idx}"].number_format = "#,##0"

        col_widths_ni = {
            "A": 6,  "B": 14, "C": 50, "D": 7,
            "E": 9,  "F": 13, "G": 11, "H": 12,
            "I": 8,  "J": 11,
        }
        for col_l, width in col_widths_ni.items():
            ws2.column_dimensions[col_l].width = width

        ws2.freeze_panes = "A3"
    else:
        ws2["A1"] = "⚠️ No se encontraron registros de tipo Non-Inventory Transfer"

    wb.save(output_path)
    return output_path


# ══════════════════════════════════════════════════════════════════════════════
#  PROCESO PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

print("\n" + "="*70)
print("PASO 1/3 — Extrayendo datos de INV TRANSFER...")
print("="*70)
inv_records, inv_meta = extract_inventory_data(PDF_PATH)
print(f"  ✅ Registros de inventario encontrados: {len(inv_records)}")
if inv_meta:
    print(f"  📄 Doc: {inv_meta.get('doc_number')}  |  Fecha: {inv_meta.get('transfer_date')}")
    print(f"  💰 Total declarado: ${inv_meta.get('total_value', 0):,.2f}")

print("\n" + "="*70)
print("PASO 2/3 — Extrayendo datos de Non-Inventory Transfer...")
print("="*70)
noninv_records, noninv_meta = extract_non_inventory_data(PDF_PATH)
print(f"  ✅ Registros non-inventory encontrados: {len(noninv_records)}")
if noninv_meta:
    print(f"  📋 Requestor: {noninv_meta.get('requestor')}")
    print(f"  💰 Total declarado: ${noninv_meta.get('total_value', 0):,.2f}")

print("\n" + "="*70)
print("PASO 3/3 — Construyendo archivo Excel...")
print("="*70)

base_name   = os.path.splitext(os.path.basename(PDF_PATH))[0]
output_path = f"{base_name}_EXTRACTED_v2.xlsx"

build_excel(inv_records, inv_meta, noninv_records, noninv_meta, output_path)
print(f"  ✅ Archivo generado: {output_path}")

# ── Estadísticas finales ──────────────────────────────────────────────────────
print("\n" + "="*70)
print("📊 RESUMEN FINAL")
print("="*70)

if inv_records:
    df_inv = pd.DataFrame(inv_records)
    df_inv["VALUE"] = df_inv["VALUE"].apply(safe_float)
    total_calc = df_inv["VALUE"].sum()
    print(f"\n  🗂️  Hoja 1 — INV TRANSFER")
    print(f"     Items extraídos : {len(inv_records)}")
    print(f"     Valor calculado : ${total_calc:,.2f}")
    total_decl = inv_meta.get("total_value", 0) or 0
    diff = total_calc - total_decl
    if total_decl:
        print(f"     Total declarado : ${total_decl:,.2f}")
        print(f"     Diferencia      : ${diff:,.2f}  {'✅ OK' if abs(diff) < 1 else '⚠️ revisar'}")

if noninv_records:
    df_ni = pd.DataFrame(noninv_records)
    df_ni["VALUE"] = df_ni["VALUE"].apply(safe_float)
    total_calc_ni = df_ni["VALUE"].sum()
    print(f"\n  🗂️  Hoja 2 — NON-INVENTORY TRANSFER")
    print(f"     Items extraídos : {len(noninv_records)}")
    print(f"     Valor calculado : ${total_calc_ni:,.2f}")
    total_decl_ni = noninv_meta.get("total_value", 0) or 0
    if total_decl_ni:
        diff_ni = total_calc_ni - total_decl_ni
        print(f"     Total declarado : ${total_decl_ni:,.2f}")
        print(f"     Diferencia      : ${diff_ni:,.2f}  {'✅ OK' if abs(diff_ni) < 1 else '⚠️ revisar'}")

print(f"\n{'='*70}")
print(f"✅ PROCESO COMPLETADO — {output_path}")
print(f"{'='*70}")

# ── Descarga automática en Colab ──────────────────────────────────────────────
try:
    from google.colab import files
    files.download(output_path)
    print(f"📥 Descargando {output_path}...")
except ImportError:
    print(f"📁 Archivo guardado en: {os.path.abspath(output_path)}")
