import streamlit as st
import pandas as pd
import pdfplumber
import re
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Validador MV",
    page_icon="🔍",
    layout="wide",
)

TOLERANCE_PESO  = 0.5
TOLERANCE_VALOR = 1.0

KEYWORDS = {
    "cantidad": ["cantidad", "piezas", "pieces", "qty", "units", "unidades"],
    "peso":     ["peso", "weight", "kg", "kilogram"],
    "valor":    ["valor", "value", "monto", "amount", "importe", "usd"],
    "factura":  ["factura", "invoice", "inv", "no factura", "num factura", "bill"],
}

# ─────────────────────────────────────────────
# CUSTOM CSS — tema claro profesional
# ─────────────────────────────────────────────
st.markdown("""
<style>
/* ── Base ── */
[data-testid="stAppViewContainer"] {
    background: linear-gradient(160deg, #f0f9ff 0%, #f8fafc 50%, #f0fdf4 100%);
    min-height: 100vh;
}
[data-testid="stHeader"] { background: transparent; }

/* ── Sidebar ── */
[data-testid="stSidebar"] {
    background: #ffffff;
    border-right: 1px solid #e2e8f0;
    box-shadow: 2px 0 12px rgba(0,0,0,.04);
}
[data-testid="stSidebar"] * { color: #334155 !important; }
[data-testid="stSidebar"] label { color: #64748b !important; font-size: 0.82rem !important; }
[data-testid="stSidebar"] [data-testid="stFileUploader"] {
    background: #f8fafc; border: 1.5px dashed #cbd5e1;
    border-radius: 10px; padding: .5rem;
}

/* ── Global text ── */
html, body, [class*="css"]  { color: #1e293b; }
h1, h2, h3, h4              { color: #0f172a !important; }

/* ── Metric cards ── */
[data-testid="metric-container"] {
    background: #ffffff;
    border: 1.5px solid #e2e8f0;
    border-radius: 14px;
    padding: 1.1rem 1.3rem;
    box-shadow: 0 2px 8px rgba(0,0,0,.05);
    transition: box-shadow .2s;
}
[data-testid="metric-container"]:hover { box-shadow: 0 4px 16px rgba(0,0,0,.09); }
[data-testid="metric-container"] label {
    color: #64748b !important; font-size: 0.76rem !important;
    text-transform: uppercase; letter-spacing: .06em; font-weight: 600 !important;
}
[data-testid="metric-container"] [data-testid="stMetricValue"] {
    color: #0f172a !important; font-size: 1.9rem !important; font-weight: 700 !important;
}
[data-testid="stMetricDelta"] { font-size: 0.77rem !important; }

/* ── DataFrames ── */
[data-testid="stDataFrame"] {
    border: 1.5px solid #e2e8f0; border-radius: 12px;
    overflow: hidden; box-shadow: 0 2px 8px rgba(0,0,0,.04);
}
iframe { border-radius: 12px; }

/* ── Expanders ── */
[data-testid="stExpander"] {
    background: #ffffff; border: 1.5px solid #e2e8f0;
    border-radius: 12px; box-shadow: 0 1px 4px rgba(0,0,0,.04);
}
[data-testid="stExpander"] summary { color: #475569 !important; font-weight: 500 !important; }

/* ── Alerts ── */
[data-testid="stAlert"] { border-radius: 12px !important; font-weight: 500; }

/* ── Download button ── */
[data-testid="stDownloadButton"] button {
    background: linear-gradient(135deg, #059669, #0d9488) !important;
    color: #ffffff !important; border: none !important;
    border-radius: 10px; font-weight: 600;
    padding: .6rem 1.6rem; transition: opacity .2s, transform .15s;
    box-shadow: 0 2px 8px rgba(5,150,105,.3);
}
[data-testid="stDownloadButton"] button:hover {
    opacity: .9 !important; transform: translateY(-1px) !important;
}

/* ── Inputs ── */
input[type="number"], input[type="text"] {
    background: #f8fafc !important; border: 1.5px solid #cbd5e1 !important;
    color: #0f172a !important; border-radius: 8px !important;
}

/* ── Divider ── */
hr { border-color: #e2e8f0 !important; margin: 1.6rem 0 !important; }

/* ── Custom components ── */
.mv-header-bar {
    background: linear-gradient(135deg, #0f766e 0%, #0369a1 100%);
    border-radius: 16px; padding: 1.5rem 2rem; margin-bottom: 1.5rem;
    box-shadow: 0 4px 20px rgba(15,118,110,.25);
}
.mv-card {
    background: #ffffff; border: 1.5px solid #e2e8f0;
    border-radius: 14px; padding: 1.2rem 1.5rem; margin-bottom: 1rem;
    box-shadow: 0 2px 8px rgba(0,0,0,.05);
}
.mv-card-title {
    font-size: .68rem; font-weight: 700; letter-spacing: .09em;
    text-transform: uppercase; color: #94a3b8; margin-bottom: .3rem;
}
.mv-card-value { font-size: 1rem; font-weight: 600; color: #0f172a; }
.section-label {
    font-size: .68rem; font-weight: 700; letter-spacing: .1em;
    text-transform: uppercase; color: #475569; margin-bottom: .5rem;
}
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
# UTILIDADES
# ─────────────────────────────────────────────

def detect_columns(df: pd.DataFrame) -> dict:
    mapping = {k: None for k in KEYWORDS}
    cols_lower = {c.lower().strip(): c for c in df.columns}
    for field, keywords in KEYWORDS.items():
        for kw in keywords:
            matches = [orig for low, orig in cols_lower.items() if kw in low]
            if matches:
                mapping[field] = matches[0]
                break
    return mapping


def parse_number(text: str):
    if not text:
        return None
    clean = re.sub(r"[^\d.,]", "", str(text)).replace(",", "")
    try:
        return float(clean)
    except ValueError:
        return None


def extract_pdf_data(pdf_file) -> dict:
    result = {"value": None, "weight": None, "pieces": None,
              "ref": None, "consignee": None, "facturas": [], "raw_text": ""}
    try:
        with pdfplumber.open(pdf_file) as pdf:
            full_text = "\n".join(p.extract_text() or "" for p in pdf.pages)
        result["raw_text"] = full_text

        m = re.search(r"Ref\s*[:\s]+(\S+)", full_text, re.IGNORECASE)
        if m:
            result["ref"] = m.group(1)

        m = re.search(r"([A-Z][A-Z ,\.]+(?:INC|LLC|SA|CORP|CO)[A-Z ,\.]*)\.",
                      full_text, re.IGNORECASE)
        if m:
            result["consignee"] = m.group(0).strip()

        m = re.search(r"([\d,\.]+)\s*[Kk][Gg]", full_text)
        if m:
            result["weight"] = parse_number(m.group(1))

        m = re.search(r"([\d,\.]+)\s*[Kk][Gg]\s+([\d,\.]+)", full_text)
        if m:
            result["pieces"] = parse_number(m.group(2))

        m = re.search(r"Value\s+([\d,\.]+)", full_text, re.IGNORECASE)
        if m:
            result["value"] = parse_number(m.group(1))

        # ── Extracción universal de facturas ──────────────────────────────────
        # Los códigos de factura siempre siguen el patrón: LETRAS-NUMEROS-LETRAS
        # (ej: RPT-0398-ENS, RIN-0085-ENS, RSCARP-0027-ENS)
        # Funciona con INV#, INV sin #, o sin prefijo ninguno.
        INVOICE_RE = re.compile(r'\b([A-Z][A-Z0-9]*-\d{3,5}-[A-Z]+)\b', re.IGNORECASE)
        found = INVOICE_RE.findall(full_text)
        # Excluir falsos positivos conocidos: refs de manifiesto tipo GSJA924294
        result["facturas"] = [
            f.upper() for f in found
            if not re.match(r'^GSJA\d+$', f, re.IGNORECASE)
        ]
    except Exception as e:
        st.error(f"Error leyendo PDF: {e}")
    return result


def semaforo(diferencia: float, tolerancia: float) -> tuple:
    a = abs(diferencia)
    if a == 0:            return "🟢", "Exacto"
    elif a <= tolerancia: return "🟡", "Redondeo"
    else:                 return "🔴", "Discrepancia"


def compare_facturas(excel_set: set, pdf_set: set) -> dict:
    en_ambos   = sorted(excel_set & pdf_set)
    solo_excel = sorted(excel_set - pdf_set)
    solo_pdf   = sorted(pdf_set   - excel_set)
    return {
        "en_ambos": en_ambos, "solo_excel": solo_excel, "solo_pdf": solo_pdf,
        "total_excel": len(excel_set), "total_pdf": len(pdf_set),
        "ok": not solo_excel and not solo_pdf,
    }


# ─────────────────────────────────────────────
# EXPORTAR EXCEL
# ─────────────────────────────────────────────

def build_excel_report(df_lines, comparison, meta, fact_result=None) -> bytes:
    wb   = Workbook()
    G    = "C6EFCE"; GF = "276221"
    Y    = "FFEB9C"; YF = "9C6500"
    R    = "FFC7CE"; RF = "9C0006"
    BH   = "0F1F3D"; W  = "FFFFFF"
    GL   = "F2F2F2"
    thin = Side(style="thin", color="BFBFBF")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(ws, row, cols):
        for c in cols:
            cell = ws.cell(row=row, column=c)
            cell.fill = PatternFill("solid", fgColor=BH)
            cell.font = Font(bold=True, color=W, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = bdr

    def styled(ws, row_i, vals, bg, fg):
        for j, v in enumerate(vals, 1):
            c = ws.cell(row_i, j)
            c.value = v
            c.fill  = PatternFill("solid", fgColor=bg)
            c.font  = Font(color=fg, size=10)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = bdr

    SC = {"Exacto": (G, GF), "Redondeo": (Y, YF), "Discrepancia": (R, RF)}

    # Hoja 1 ─ Comparación
    ws1 = wb.active; ws1.title = "Comparación"
    ws1.row_dimensions[1].height = 30
    ws1.merge_cells("A1:H1")
    ws1["A1"] = "REPORTE — VALIDADOR MANIFESTACIÓN DE VALOR vs ACEM"
    ws1["A1"].font = Font(bold=True, color=W, size=13)
    ws1["A1"].fill = PatternFill("solid", fgColor=BH)
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for r, (lbl, val) in enumerate([
        ("Referencia", meta.get("ref", "—")), ("Fecha", meta.get("fecha", "—")),
        ("Consignatario", meta.get("consignee", "—")),
        ("Excel", meta.get("excel_name", "—")), ("PDF", meta.get("pdf_name", "—")),
    ], 2):
        ws1.cell(r, 1).value = lbl; ws1.cell(r, 1).font = Font(bold=True, size=10)
        ws1.cell(r, 2).value = val; ws1.cell(r, 2).font = Font(size=10)

    HDRS = ["Campo", "Etiqueta Excel", "Etiqueta PDF",
            "Valor Excel", "Valor PDF", "Tolerancia", "Estado", "Observaciones"]
    hdr(ws1, 8, range(1, len(HDRS) + 1))
    for i, h in enumerate(HDRS, 1):
        ws1.cell(8, i).value = h
    for i, row in enumerate(comparison, 9):
        bg, fg = SC.get(row["estado"], (GL, "000000"))
        styled(ws1, i, [row["campo"], row["etiqueta_excel"], row["etiqueta_pdf"],
                        row["val_excel"], row["val_pdf"], row["tolerancia"],
                        f"{row['emoji']} {row['estado']}", row["obs"]], bg, fg)
    for i, w in enumerate([22,18,18,14,14,13,16,44], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # Hoja 2 ─ Desglose
    ws2 = wb.create_sheet("Desglose Excel")
    ws2.merge_cells("A1:H1")
    ws2["A1"] = "DESGLOSE DE LÍNEAS — ARCHIVO EXCEL"
    ws2["A1"].font = Font(bold=True, color=W, size=12)
    ws2["A1"].fill = PatternFill("solid", fgColor=BH)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 25
    if not df_lines.empty:
        cols = list(df_lines.columns)
        for j, col in enumerate(cols, 1): ws2.cell(2, j).value = col
        hdr(ws2, 2, range(1, len(cols)+1))
        for i, (_, r) in enumerate(df_lines.iterrows(), 3):
            bg = GL if i % 2 == 1 else W
            for j, v in enumerate(r.values, 1):
                c = ws2.cell(i, j)
                c.value = v; c.fill = PatternFill("solid", fgColor=bg)
                c.font = Font(size=10); c.border = bdr
                c.alignment = Alignment(horizontal="center", vertical="center")
        tot = len(df_lines) + 3
        ws2.cell(tot, 1).value = "TOTAL"
        ws2.cell(tot, 1).font  = Font(bold=True, color=W, size=10)
        ws2.cell(tot, 1).fill  = PatternFill("solid", fgColor=BH); ws2.cell(tot, 1).border = bdr
        for j, col in enumerate(cols, 1):
            if pd.api.types.is_numeric_dtype(df_lines[col]):
                c = ws2.cell(tot, j)
                c.value = f"=SUM({get_column_letter(j)}3:{get_column_letter(j)}{tot-1})"
                c.font  = Font(bold=True, color=W, size=10)
                c.fill  = PatternFill("solid", fgColor=BH); c.border = bdr
                c.alignment = Alignment(horizontal="center")
            elif j > 1:
                ws2.cell(tot, j).fill = PatternFill("solid", fgColor=BH)
                ws2.cell(tot, j).border = bdr
        for j in range(1, len(cols)+1):
            ws2.column_dimensions[get_column_letter(j)].width = 22

    # Hoja 3 ─ Diagnóstico + Facturas
    ws3 = wb.create_sheet("Diagnóstico")
    ws3.merge_cells("A1:D1")
    ws3["A1"] = "DIAGNÓSTICO Y VALIDACIÓN DE FACTURAS"
    ws3["A1"].font = Font(bold=True, color=W, size=12)
    ws3["A1"].fill = PatternFill("solid", fgColor=BH)
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 25

    exact  = sum(1 for r in comparison if r["estado"] == "Exacto")
    redond = sum(1 for r in comparison if r["estado"] == "Redondeo")
    discr  = sum(1 for r in comparison if r["estado"] == "Discrepancia")
    diag   = [
        ("Total campos comparados", len(comparison)),
        ("Coincidencias exactas 🟢", exact),
        ("Diferencias por redondeo 🟡", redond),
        ("Discrepancias críticas 🔴", discr),
        ("Resultado global", "✅ SIN ERRORES" if discr == 0 else "⚠️ REVISAR"),
    ]
    for i, (lbl, val) in enumerate(diag, 3):
        bg = GL if i % 2 == 0 else W
        ws3.cell(i, 1).value = lbl; ws3.cell(i, 1).font = Font(bold=True, size=11)
        ws3.cell(i, 2).value = val; ws3.cell(i, 2).font = Font(size=11)
        for j in [1, 2]:
            ws3.cell(i, j).fill = PatternFill("solid", fgColor=bg)
            ws3.cell(i, j).border = bdr
            ws3.cell(i, j).alignment = Alignment(vertical="center", horizontal="center")
        ws3.row_dimensions[i].height = 20

    if fact_result:
        start = len(diag) + 5
        for j in [1, 2]:
            ws3.cell(start, j).fill = PatternFill("solid", fgColor=BH); ws3.cell(start, j).border = bdr
        ws3.cell(start, 1).value = "VALIDACIÓN DE FACTURAS (INV#)"
        ws3.cell(start, 1).font  = Font(bold=True, color=W, size=11)
        ws3.row_dimensions[start].height = 20
        hdr(ws3, start+1, range(1, 5))
        for j, h in enumerate(["Factura","En Excel","En PDF (INV#)","Estado"], 1):
            ws3.cell(start+1, j).value = h
        all_inv = sorted(
            set(fact_result.get("en_ambos",[])) |
            set(fact_result.get("solo_excel",[])) |
            set(fact_result.get("solo_pdf",[]))
        )
        for idx, inv in enumerate(all_inv, start+2):
            en_xl  = inv in fact_result.get("en_ambos",[]) + fact_result.get("solo_excel",[])
            en_pdf = inv in fact_result.get("en_ambos",[]) + fact_result.get("solo_pdf",[])
            if en_xl and en_pdf:  sf, bg_f, fg_f = "✅ Coincide",     G, GF
            elif en_xl:           sf, bg_f, fg_f = "🔴 Solo Excel",   R, RF
            else:                 sf, bg_f, fg_f = "🔴 Solo PDF",     R, RF
            styled(ws3, idx, [inv, "✔" if en_xl else "✘", "✔" if en_pdf else "✘", sf], bg_f, fg_f)

    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 35

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ─────────────────────────────────────────────
# UI PRINCIPAL
# ─────────────────────────────────────────────

def main():
    st.markdown("""
    <div class="mv-header-bar">
      <h1 style="margin:0;font-size:1.6rem;font-weight:700;color:#ffffff;letter-spacing:-.01em">
        🔍 Validador Manifestación de Valor
      </h1>
      <p style="margin:.3rem 0 0 0;color:rgba(255,255,255,.72);font-size:.88rem">
        Comparación automática de valores entre el archivo Excel del SEER y el Manifiesto ACEM
      </p>
    </div>
    """, unsafe_allow_html=True)
    st.divider()

    # ── Sidebar ──
    with st.sidebar:
        st.markdown('<p class="section-label" style="margin-top:.4rem;color:#94a3b8">Archivos de entrada</p>',
                    unsafe_allow_html=True)
        excel_file = st.file_uploader("📊 Excel — CCP / EFO", type=["xlsx","xls"])
        pdf_file   = st.file_uploader("📄 PDF — Manifiesto ACEM", type=["pdf"])
        st.divider()
        st.markdown('<p class="section-label" style="color:#94a3b8">Tolerancias de validación</p>', unsafe_allow_html=True)
        tol_peso  = st.number_input("Peso (Kg)",   value=TOLERANCE_PESO,  step=0.1, format="%.2f")
        tol_valor = st.number_input("Valor (USD)",  value=TOLERANCE_VALOR, step=0.1, format="%.2f")

    if not excel_file or not pdf_file:
        st.markdown("""
        <div style="background:#ffffff;border:1.5px dashed #cbd5e1;border-radius:16px;
                    padding:3.5rem 2rem;text-align:center;margin-top:1rem;
                    box-shadow:0 2px 8px rgba(0,0,0,.04)">
          <p style="font-size:2.5rem;margin:0">📂</p>
          <p style="color:#64748b;margin:.7rem 0 .3rem 0;font-size:1rem;font-weight:500">
            Sube los archivos en el panel izquierdo
          </p>
          <p style="color:#94a3b8;font-size:.84rem;margin:0">
            Excel (CCP / EFO) · PDF (Manifiesto ACEM)
          </p>
        </div>
        """, unsafe_allow_html=True)
        return

    # ── Leer archivos ──
    try:
        df = pd.read_excel(excel_file)
    except Exception as e:
        st.error(f"No se pudo leer el Excel: {e}"); return

    col_map  = detect_columns(df)
    pdf_data = extract_pdf_data(pdf_file)

    missing = [k for k, v in col_map.items() if v is None]
    if missing:
        with st.expander("⚠️ Columnas no detectadas — selecciónalas manualmente"):
            cols_avail = ["(ninguna)"] + list(df.columns)
            for field in missing:
                sel = st.selectbox(f"Columna para '{field}'", cols_avail)
                if sel != "(ninguna)":
                    col_map[field] = sel

    # ── Totales ──
    def safe_sum(col):
        if col and col in df.columns:
            return pd.to_numeric(df[col], errors="coerce").sum()
        return None

    ex_cant  = safe_sum(col_map["cantidad"])
    ex_peso  = safe_sum(col_map["peso"])
    ex_valor = safe_sum(col_map["valor"])

    # ── Comparación numérica ──
    comparison = []
    for campo, etiq_xl, etiq_pdf, val_xl, val_pdf, tol, unidad in [
        ("Cantidad / Piezas", col_map["cantidad"] or "—", "Pieces",      ex_cant,  pdf_data["pieces"], 0,         "pcs"),
        ("Peso bruto",        col_map["peso"]     or "—", "Weight (Kg)", ex_peso,  pdf_data["weight"], tol_peso,  "Kg"),
        ("Valor mercancía",   col_map["valor"]    or "—", "Value",       ex_valor, pdf_data["value"],  tol_valor, "USD"),
    ]:
        if val_xl is None or val_pdf is None:
            emoji, estado, obs = "⚪", "Sin dato", "Valor no encontrado en uno de los documentos."
        else:
            diff = val_xl - val_pdf
            emoji, estado = semaforo(diff, tol)
            obs = (
                "Coincidencia exacta." if estado == "Exacto"
                else f"Diferencia de {round(diff,4)} {unidad} por redondeo." if estado == "Redondeo"
                else f"Discrepancia de {round(diff,4)} {unidad}. Verificar con agente."
            )
        comparison.append({
            "campo": campo, "etiqueta_excel": etiq_xl, "etiqueta_pdf": etiq_pdf,
            "val_excel": round(val_xl, 4) if isinstance(val_xl, float) else val_xl,
            "val_pdf":   round(val_pdf, 4) if isinstance(val_pdf, float) else val_pdf,
            "tolerancia": f"±{tol} {unidad}",
            "emoji": emoji, "estado": estado, "obs": obs,
        })

    # ── Facturas ──
    col_factura = col_map.get("factura")
    excel_facturas = (
        set(df[col_factura].dropna().astype(str).str.strip().unique())
        if col_factura and col_factura in df.columns else set()
    )
    pdf_facturas = set(pdf_data.get("facturas", []))
    fact_result  = compare_facturas(excel_facturas, pdf_facturas)

    exact_c  = sum(1 for r in comparison if r["estado"] == "Exacto")
    redond_c = sum(1 for r in comparison if r["estado"] == "Redondeo")
    discr_c  = sum(1 for r in comparison if r["estado"] == "Discrepancia")

    # ── Tarjeta de referencia ──
    ref       = pdf_data.get("ref") or excel_file.name
    consignee = pdf_data.get("consignee", "—")
    st.markdown(f"""
    <div class="mv-card" style="display:flex;gap:2.5rem;flex-wrap:wrap;align-items:center">
      <div>
        <div class="mv-card-title">Referencia</div>
        <div class="mv-card-value" style="color:#0f766e;font-size:1.3rem;font-weight:700">{ref}</div>
      </div>
      <div style="width:1px;background:#e2e8f0;height:40px;align-self:center"></div>
      <div>
        <div class="mv-card-title">Consignatario</div>
        <div class="mv-card-value">{consignee}</div>
      </div>
      <div style="width:1px;background:#e2e8f0;height:40px;align-self:center"></div>
      <div>
        <div class="mv-card-title">Excel</div>
        <div class="mv-card-value" style="font-size:.85rem;color:#64748b">{excel_file.name}</div>
      </div>
      <div>
        <div class="mv-card-title">PDF</div>
        <div class="mv-card-value" style="font-size:.85rem;color:#64748b">{pdf_file.name}</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Métricas ──
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Campos comparados", len(comparison))
    c2.metric("🟢 Exactos",   exact_c)
    c3.metric("🟡 Redondeo",  redond_c)
    c4.metric("🔴 Discrepancias", discr_c,
              delta="REVISAR" if discr_c > 0 else None, delta_color="inverse")
    max_f = max(fact_result["total_excel"], fact_result["total_pdf"], 1)
    c5.metric("🧾 Facturas",
              f"{len(fact_result['en_ambos'])}/{max_f}",
              delta=None if fact_result["ok"] else "Sin coincidencia",
              delta_color="inverse")

    st.divider()

    # ── Banner resultado global ──
    if discr_c == 0 and redond_c == 0 and fact_result["ok"]:
        st.success("✅ Todos los valores y facturas coinciden exactamente.")
    elif discr_c == 0 and fact_result["ok"]:
        st.warning("🟡 Documentos consistentes. Diferencias menores por redondeo.")
    else:
        st.error("🔴 Se detectaron discrepancias. Revisa el detalle a continuación.")

    st.divider()

    # ══════════════════════════
    # TABLA COMPARATIVA
    # ══════════════════════════
    st.markdown('<p class="section-label">Tabla comparativa — valores numéricos</p>',
                unsafe_allow_html=True)

    df_cmp = pd.DataFrame([{
        "Campo":          r["campo"],
        "Etiqueta Excel": r["etiqueta_excel"],
        "Etiqueta PDF":   r["etiqueta_pdf"],
        "Valor Excel":    r["val_excel"],
        "Valor PDF":      r["val_pdf"],
        "Tolerancia":     r["tolerancia"],
        "Estado":         f"{r['emoji']} {r['estado']}",
        "Observaciones":  r["obs"],
    } for r in comparison])

    def hl_estado(row):
        e = row["Estado"]
        if "Exacto"       in e: return ["background-color:#dcfce7;color:#166534"] * len(row)
        if "Redondeo"     in e: return ["background-color:#fef9c3;color:#854d0e"] * len(row)
        if "Discrepancia" in e: return ["background-color:#fee2e2;color:#991b1b"] * len(row)
        return ["background-color:#f8fafc;color:#64748b"] * len(row)

    st.dataframe(
        df_cmp.style.apply(hl_estado, axis=1),
        use_container_width=True, hide_index=True,
    )

    st.divider()

    # ══════════════════════════
    # VALIDACIÓN FACTURAS
    # ══════════════════════════
    st.markdown('<p class="section-label">Validación de facturas — INV#</p>',
                unsafe_allow_html=True)

    all_inv = sorted(excel_facturas | pdf_facturas)
    if all_inv:
        if fact_result["ok"]:
            invs_str = "  ·  ".join(fact_result["en_ambos"])
            st.markdown(f"""
            <div style="background:#f0fdf4;border:1.5px solid #86efac;border-radius:12px;
                        padding:.9rem 1.3rem;display:flex;align-items:center;
                        gap:.8rem;margin-bottom:.8rem;box-shadow:0 1px 4px rgba(22,163,74,.1)">
              <span style="font-size:1.3rem">✅</span>
              <div>
                <span style="color:#166534;font-weight:700;font-size:.95rem">Facturas coinciden</span>
                <span style="color:#15803d;font-size:.85rem;margin-left:.8rem">{invs_str}</span>
              </div>
            </div>
            """, unsafe_allow_html=True)
        else:
            msgs = []
            if fact_result["solo_excel"]:
                msgs.append(f"Solo en Excel: <strong>{', '.join(fact_result['solo_excel'])}</strong>")
            if fact_result["solo_pdf"]:
                msgs.append(f"Solo en PDF: <strong>{', '.join(fact_result['solo_pdf'])}</strong>")
            st.markdown(f"""
            <div style="background:#fef2f2;border:1.5px solid #fca5a5;border-radius:12px;
                        padding:.9rem 1.3rem;display:flex;align-items:center;
                        gap:.8rem;margin-bottom:.8rem;box-shadow:0 1px 4px rgba(220,38,38,.1)">
              <span style="font-size:1.3rem">🔴</span>
              <span style="color:#991b1b;font-size:.9rem">{' &nbsp;|&nbsp; '.join(msgs)}</span>
            </div>
            """, unsafe_allow_html=True)

        rows_f = []
        for inv in all_inv:
            en_xl  = inv in excel_facturas
            en_pdf = inv in pdf_facturas
            if en_xl and en_pdf:   ef, emj = "Coincide",     "🟢"
            elif en_xl:            ef, emj = "Solo en Excel", "🔴"
            else:                  ef, emj = "Solo en PDF",   "🔴"
            rows_f.append({"Factura": inv,
                           "En Excel":      "✔" if en_xl  else "✘",
                           "En PDF (INV#)": "✔" if en_pdf else "✘",
                           "Estado":        f"{emj} {ef}"})

        def hl_fact(row):
            if "Coincide" in row["Estado"]:
                return ["background-color:#dcfce7;color:#166534"] * len(row)
            return ["background-color:#fee2e2;color:#991b1b"] * len(row)

        st.dataframe(
            pd.DataFrame(rows_f).style.apply(hl_fact, axis=1),
            use_container_width=True, hide_index=True,
        )
        st.caption(
            f"Columna Excel: **{col_factura}**  ·  "
            f"Facturas extraídas del PDF (INV#): {', '.join(sorted(pdf_facturas)) or '(ninguna)'}"
            if col_factura else
            "⚠️ Columna de factura no detectada automáticamente."
        )
    else:
        st.info("No hay facturas para comparar. Verifica que el PDF contenga una línea INV#.")

    st.divider()

    # ══════════════════════════
    # EXPANDERS
    # ══════════════════════════
    with st.expander("📄 Desglose de líneas del Excel"):
        st.dataframe(df, use_container_width=True, hide_index=True)

    with st.expander("🔎 Texto extraído del PDF"):
        st.text(pdf_data["raw_text"] or "(sin texto extraído)")

    with st.expander("🔧 Columnas detectadas automáticamente"):
        st.dataframe(
            pd.DataFrame([{"Campo": k, "Columna detectada": v or "⚠️ No detectada"}
                          for k, v in col_map.items()]),
            use_container_width=True, hide_index=True,
        )

    st.divider()

    # ── Exportar ──
    meta = {
        "ref": ref, "fecha": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "consignee": consignee, "excel_name": excel_file.name, "pdf_name": pdf_file.name,
    }
    excel_bytes = build_excel_report(df, comparison, meta, fact_result)
    st.download_button(
        label="⬇️  Descargar reporte en Excel",
        data=excel_bytes,
        file_name=f"validacion_{ref}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    main()
